from __future__ import annotations

import threading
from io import BytesIO
from zipfile import ZipFile

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image

from fashion_cms.image_downloader import (
    MIB,
    REPORT_HEADERS,
    DownloadSettings,
    build_download_report,
    build_image_zip,
    download_images,
    parse_url_workbook,
)
from fashion_cms.models import DownloadResult, ImageDownloadResult, ImageUrlRequest


PUBLIC_IP = "93.184.216.34"
PUBLIC_URL = "https://images.example/item.jpg"


def public_resolver(_host: str, _port: int) -> tuple[str, ...]:
    return (PUBLIC_IP,)


def image_bytes(
    image_format: str = "JPEG",
    size: tuple[int, int] = (40, 20),
    mode: str = "RGB",
    color: object = "blue",
    *,
    orientation: int | None = None,
) -> bytes:
    image = Image.new(mode, size, color)
    output = BytesIO()
    exif = Image.Exif()
    if orientation is not None:
        exif[274] = orientation
    image.save(output, format=image_format, exif=exif)
    image.close()
    return output.getvalue()


def workbook_bytes(*rows: tuple[object, ...]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    width = max(map(len, rows), default=3)
    worksheet.append(["SKU", *(f"URL {ordinal}" for ordinal in range(1, width))])
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def request(url: str = PUBLIC_URL, *, sku: str = "SKU", ordinal: int = 1) -> ImageUrlRequest:
    return ImageUrlRequest(row_number=2, sku=sku, ordinal=ordinal, source_url=url)


def mocked_download(
    requests: tuple[ImageUrlRequest, ...],
    handler,
    *,
    settings: DownloadSettings | None = None,
    resolver=public_resolver,
    previous: ImageDownloadResult | None = None,
) -> ImageDownloadResult:
    return download_images(
        requests,
        settings=settings,
        transport=httpx.MockTransport(handler),
        resolver=resolver,
        sleeper=lambda _delay: None,
        jitter=lambda: 0,
        previous=previous,
    )


def image_response(content: bytes, content_type: str = "image/jpeg") -> httpx.Response:
    return httpx.Response(200, headers={"Content-Type": content_type}, content=content)


def test_successful_jpeg_download_uses_required_name_and_dimensions() -> None:
    result = mocked_download((request(),), lambda _request: image_response(image_bytes()))

    assert len(result.images) == 1
    image = result.images[0]
    assert (image.output_filename, image.source_width, image.source_height) == (
        "SKU-1.jpg",
        40,
        20,
    )
    assert result.report[0].model_dump() == {
        "sku": "SKU",
        "ordinal": 1,
        "source_url": PUBLIC_URL,
        "result": DownloadResult.SUCCESS,
        "http_status": 200,
        "output_filename": "SKU-1.jpg",
        "source_dimensions": (40, 20),
        "output_dimensions": (1500, 1500),
        "error_message": None,
    }
    with Image.open(BytesIO(image.content)) as output:
        assert (output.format, output.mode, output.size) == ("JPEG", "RGB", (1500, 1500))


def test_transparent_png_is_composited_onto_white() -> None:
    source = Image.new("RGBA", (200, 100), (0, 0, 0, 0))
    source.paste((255, 0, 0, 255), (0, 0, 100, 100))
    content = BytesIO()
    source.save(content, format="PNG")
    source.close()

    result = mocked_download(
        (request(),),
        lambda _request: image_response(content.getvalue(), "image/png"),
    )

    with Image.open(BytesIO(result.images[0].content)) as output:
        red = output.getpixel((700, 750))
        white = output.getpixel((800, 750))
        assert red[0] > 240 and red[1] < 20 and red[2] < 20
        assert min(white) > 240


def test_cmyk_jpeg_is_converted_safely_to_rgb() -> None:
    content = image_bytes(mode="CMYK", color=(0, 255, 255, 0))
    result = mocked_download((request(),), lambda _request: image_response(content))

    with Image.open(BytesIO(result.images[0].content)) as output:
        red, green, blue = output.getpixel((750, 750))
        assert output.mode == "RGB"
        assert red > 200 and green < 60 and blue < 60


def test_exif_orientation_is_applied_before_reporting_dimensions() -> None:
    content = image_bytes(size=(40, 20), orientation=6)
    result = mocked_download((request(),), lambda _request: image_response(content))

    assert result.report[0].source_dimensions == (20, 40)
    assert (result.images[0].source_width, result.images[0].source_height) == (20, 40)


def test_redirect_to_public_url_is_revalidated_and_downloaded() -> None:
    visited: list[tuple[str, str]] = []

    def handler(incoming: httpx.Request) -> httpx.Response:
        visited.append((str(incoming.url), incoming.headers["host"]))
        if incoming.url.path == "/start":
            return httpx.Response(302, headers={"Location": PUBLIC_URL})
        return image_response(image_bytes())

    result = mocked_download((request("https://redirect.example/start"),), handler)

    assert [row.result for row in result.report] == [DownloadResult.SUCCESS]
    assert visited == [
        (f"https://{PUBLIC_IP}/start", "redirect.example"),
        (f"https://{PUBLIC_IP}/item.jpg", "images.example"),
    ]


def test_redirect_to_private_url_is_rejected_without_following_it() -> None:
    visited: list[tuple[str, str]] = []

    def handler(incoming: httpx.Request) -> httpx.Response:
        visited.append((str(incoming.url), incoming.headers["host"]))
        return httpx.Response(302, headers={"Location": "http://127.0.0.1/secret"})

    result = mocked_download((request("https://redirect.example/start"),), handler)

    assert result.images == ()
    assert result.report[0].http_status == 302
    assert "not allowed" in (result.report[0].error_message or "")
    assert visited == [(f"https://{PUBLIC_IP}/start", "redirect.example")]


def test_connection_is_pinned_to_validated_ip_with_original_host_and_sni() -> None:
    observed: dict[str, object] = {}

    def handler(incoming: httpx.Request) -> httpx.Response:
        observed.update(
            url=str(incoming.url),
            host=incoming.headers["host"],
            sni=incoming.extensions["sni_hostname"],
        )
        return image_response(image_bytes())

    result = mocked_download((request(),), handler)

    assert result.images
    assert observed == {
        "url": f"https://{PUBLIC_IP}/item.jpg",
        "host": "images.example",
        "sni": "images.example",
    }


@pytest.mark.parametrize(
    "url",
    [
        "http://127.0.0.1/image.jpg",
        "http://localhost/image.jpg",
        "http://10.0.0.1/image.jpg",
        "http://169.254.169.254/latest/meta-data",
        "http://[fec0::1]/image.jpg",
        "http://metadata.google.internal/computeMetadata/v1/",
        "file:///tmp/image.jpg",
    ],
)
def test_direct_private_local_and_non_http_urls_are_rejected_before_request(url: str) -> None:
    calls = 0

    def handler(_request: httpx.Request) -> httpx.Response:
        nonlocal calls
        calls += 1
        return image_response(image_bytes())

    result = mocked_download((request(url),), handler)

    assert calls == 0
    assert result.images == ()
    assert result.report[0].result == DownloadResult.FAILED


def test_hostname_resolving_to_private_ip_is_rejected_before_request() -> None:
    result = mocked_download(
        (request("https://internal.example/image.jpg"),),
        lambda _request: pytest.fail("private DNS result must not be fetched"),
        resolver=lambda _host, _port: ("10.0.0.1",),
    )

    assert result.images == ()
    assert "non-public" in (result.report[0].error_message or "")


def test_timeout_is_retried_three_times_then_reported() -> None:
    calls = 0

    def handler(incoming: httpx.Request) -> httpx.Response:
        nonlocal calls
        calls += 1
        raise httpx.ReadTimeout("timed out", request=incoming)

    result = mocked_download((request(),), handler)

    assert calls == 4
    assert result.images == ()
    assert result.report[0].error_message == "Request timed out."


@pytest.mark.parametrize(
    ("status", "expected_calls"),
    [(403, 1), (429, 4), (500, 4)],
)
def test_http_errors_have_expected_retry_policy(status: int, expected_calls: int) -> None:
    calls = 0

    def handler(_request: httpx.Request) -> httpx.Response:
        nonlocal calls
        calls += 1
        return httpx.Response(status)

    result = mocked_download((request(),), handler)

    assert calls == expected_calls
    assert result.images == ()
    assert result.report[0].http_status == status
    assert result.report[0].error_message == f"HTTP {status} response."


def test_retry_backoff_is_bounded_exponential_with_deterministic_jitter() -> None:
    delays: list[float] = []
    calls = 0

    def handler(_request: httpx.Request) -> httpx.Response:
        nonlocal calls
        calls += 1
        return httpx.Response(500)

    result = download_images(
        (request(),),
        transport=httpx.MockTransport(handler),
        resolver=public_resolver,
        sleeper=delays.append,
        jitter=lambda: 0,
    )

    assert result.images == ()
    assert calls == 4
    assert delays == [0.25, 0.5, 1.0]


@pytest.mark.parametrize("declared_size", [26 * MIB, -1])
def test_invalid_or_oversized_declared_response_is_rejected(declared_size: int) -> None:
    result = mocked_download(
        (request(),),
        lambda _request: httpx.Response(
            200,
            headers={
                "Content-Type": "image/jpeg",
                "Content-Length": str(declared_size),
            },
            content=b"x",
        ),
        settings=DownloadSettings(retry_count=0),
    )

    assert result.images == ()
    assert "limit" in (result.report[0].error_message or "")


class UndeclaredStream(httpx.SyncByteStream):
    def __iter__(self):
        yield b"123456"
        yield b"789012"


def test_streamed_response_size_is_limited_without_content_length() -> None:
    result = mocked_download(
        (request(),),
        lambda _request: httpx.Response(
            200,
            headers={"Content-Type": "image/jpeg"},
            stream=UndeclaredStream(),
        ),
        settings=DownloadSettings(max_response_bytes=10, retry_count=0),
    )

    assert result.images == ()
    assert "limit" in (result.report[0].error_message or "")


def test_decoded_image_pixel_limit_is_enforced() -> None:
    result = mocked_download(
        (request(),),
        lambda _request: image_response(image_bytes()),
        settings=DownloadSettings(max_decoded_pixels=100, retry_count=0),
    )

    assert result.images == ()
    assert result.report[0].error_message == "Image exceeds 100 decoded pixels."


def test_html_content_type_is_rejected_even_with_valid_image_bytes() -> None:
    result = mocked_download(
        (request(),),
        lambda _request: image_response(image_bytes(), "text/html"),
        settings=DownloadSettings(retry_count=0),
    )

    assert result.images == ()
    assert result.report[0].error_message == (
        "Response Content-Type is not a supported image type."
    )


@pytest.mark.parametrize("content", [b"<html>not an image</html>", b"broken image bytes"])
def test_image_content_type_with_non_image_bytes_is_rejected(content: bytes) -> None:
    result = mocked_download(
        (request(),),
        lambda _request: image_response(content),
        settings=DownloadSettings(retry_count=0),
    )

    assert result.images == ()
    assert "decode" in (result.report[0].error_message or "").casefold()


def test_blank_url_one_keeps_valid_url_two_at_ordinal_two() -> None:
    parsed = parse_url_workbook(workbook_bytes(("SKU", None, PUBLIC_URL)))

    assert parsed.ready
    assert [(item.sku, item.ordinal, item.source_url) for item in parsed.requests] == [
        ("SKU", 2, PUBLIC_URL)
    ]
    result = mocked_download(parsed.requests, lambda _request: image_response(image_bytes()))
    assert [image.output_filename for image in result.images] == ["SKU-2.jpg"]


def test_failed_url_one_does_not_renumber_successful_url_two() -> None:
    failed_url = "https://images.example/forbidden.jpg"
    parsed = parse_url_workbook(workbook_bytes(("SKU", failed_url, PUBLIC_URL)))

    def handler(incoming: httpx.Request) -> httpx.Response:
        if incoming.url.path == "/forbidden.jpg":
            return httpx.Response(403)
        return image_response(image_bytes())

    result = mocked_download(parsed.requests, handler)

    assert [image.output_filename for image in result.images] == ["SKU-2.jpg"]
    assert [(row.ordinal, row.result) for row in result.report] == [
        (1, DownloadResult.FAILED),
        (2, DownloadResult.SUCCESS),
    ]


def test_leading_zero_sku_remains_text_through_download() -> None:
    parsed = parse_url_workbook(workbook_bytes(("00123", PUBLIC_URL)))

    assert parsed.ready
    assert parsed.requests[0].sku == "00123"
    result = mocked_download(parsed.requests, lambda _request: image_response(image_bytes()))
    assert result.images[0].output_filename == "00123-1.jpg"


def test_pad_white_preserves_aspect_ratio_and_does_not_upscale_small_images() -> None:
    small = image_bytes("PNG", (200, 100))
    small_result = mocked_download(
        (request(),), lambda _request: image_response(small, "image/png")
    )
    large = image_bytes("PNG", (2800, 700))
    large_result = mocked_download(
        (request(),), lambda _request: image_response(large, "image/png")
    )

    assert small_result.images[0].low_resolution
    with Image.open(BytesIO(small_result.images[0].content)) as output:
        assert (output.mode, output.size) == ("RGB", (1500, 1500))
        assert output.getpixel((700, 750))[2] > 240
        assert min(output.getpixel((600, 750))) > 240
        assert min(output.getpixel((750, 650))) > 240
    with Image.open(BytesIO(large_result.images[0].content)) as output:
        assert output.getpixel((60, 750))[2] > 240
        assert output.getpixel((1440, 750))[2] > 240
        assert min(output.getpixel((750, 550))) > 240


def test_zip_is_flat_success_only_sorted_and_byte_deterministic() -> None:
    requests = (
        request("https://images.example/z.jpg", sku="Z", ordinal=2),
        request("https://images.example/a.jpg", sku="A", ordinal=1),
        request("https://images.example/fail.jpg", sku="B", ordinal=1),
    )

    def handler(incoming: httpx.Request) -> httpx.Response:
        if incoming.url.path == "/fail.jpg":
            return httpx.Response(403)
        return image_response(image_bytes())

    result = mocked_download(requests, handler)
    first = build_image_zip(result.images)
    second = build_image_zip(tuple(reversed(result.images)))

    assert first == second
    with ZipFile(BytesIO(first)) as archive:
        assert archive.namelist() == ["A-1.jpg", "Z-2.jpg"]
        assert all("/" not in name and "\\" not in name for name in archive.namelist())
        assert all(info.date_time == (1980, 1, 1, 0, 0, 0) for info in archive.infolist())


def test_download_report_is_separate_and_has_exact_fields_for_failure() -> None:
    requests = (
        request("https://images.example/ok.jpg", sku="A"),
        request("https://images.example/forbidden.jpg", sku="B"),
    )

    def handler(incoming: httpx.Request) -> httpx.Response:
        if incoming.url.path == "/forbidden.jpg":
            return httpx.Response(403)
        return image_response(image_bytes())

    result = mocked_download(requests, handler)
    report_content = build_download_report(result.report)
    workbook = load_workbook(BytesIO(report_content), read_only=True, data_only=False)
    rows = list(workbook.active.iter_rows(values_only=True))
    workbook.close()

    assert rows[0] == REPORT_HEADERS
    assert rows[1] == (
        "A",
        1,
        "https://images.example/ok.jpg",
        "SUCCESS",
        200,
        "A-1.jpg",
        "40 × 20",
        "1500 × 1500",
        None,
    )
    assert rows[2] == (
        "B",
        1,
        "https://images.example/forbidden.jpg",
        "FAILED",
        403,
        None,
        None,
        None,
        "HTTP 403 response.",
    )


def test_untrusted_report_text_is_literal_and_xml_safe() -> None:
    hostile = request("https://images.example/a\x01b", sku="=SKU")
    result = mocked_download(
        (hostile,),
        lambda _request: pytest.fail("control-character URL must not be fetched"),
    )

    workbook = load_workbook(BytesIO(build_download_report(result.report)), data_only=False)
    worksheet = workbook.active
    assert worksheet["A2"].value == "=SKU"
    assert worksheet["A2"].data_type == "s"
    assert worksheet["C2"].value == "https://images.example/a�b"
    assert worksheet["C2"].data_type == "s"
    workbook.close()


def test_retry_failures_reuses_success_without_refetching_it() -> None:
    ok = request("https://images.example/ok.jpg", sku="OK")
    retry = request("https://images.example/retry.jpg", sku="RETRY")
    calls = {"/ok.jpg": 0, "/retry.jpg": 0}
    retry_succeeds = False

    def handler(incoming: httpx.Request) -> httpx.Response:
        nonlocal retry_succeeds
        path = incoming.url.path
        calls[path] += 1
        if path == "/retry.jpg" and not retry_succeeds:
            return httpx.Response(500)
        return image_response(image_bytes())

    first = mocked_download((ok, retry), handler)
    assert [image.sku for image in first.images] == ["OK"]
    assert calls == {"/ok.jpg": 1, "/retry.jpg": 4}

    retry_succeeds = True
    second = mocked_download((ok, retry), handler, previous=first)

    assert [image.sku for image in second.images] == ["OK", "RETRY"]
    assert calls == {"/ok.jpg": 1, "/retry.jpg": 5}
    assert all(row.result == DownloadResult.SUCCESS for row in second.report)


def test_default_limits_timeouts_and_retry_count_are_applied() -> None:
    settings = DownloadSettings()
    observed_timeout: dict[str, float] = {}

    def handler(incoming: httpx.Request) -> httpx.Response:
        observed_timeout.update(incoming.extensions["timeout"])
        return image_response(image_bytes())

    result = mocked_download((request(),), handler)

    assert result.images
    assert settings == DownloadSettings(
        total_concurrency=8,
        per_host_concurrency=4,
        connect_timeout_seconds=10,
        read_timeout_seconds=30,
        retry_count=3,
        max_response_bytes=25 * MIB,
        max_decoded_pixels=50_000_000,
    )
    assert observed_timeout == {"connect": 10.0, "read": 30.0, "write": 10.0, "pool": 10.0}


def test_total_and_per_host_concurrency_defaults_are_enforced() -> None:
    gate = threading.Event()
    lock = threading.Lock()
    active = 0
    peak_total = 0
    peak_by_host: dict[str, int] = {}
    active_by_host: dict[str, int] = {}
    content = image_bytes()

    def handler(incoming: httpx.Request) -> httpx.Response:
        nonlocal active, peak_total
        host = incoming.headers["host"]
        with lock:
            active += 1
            active_by_host[host] = active_by_host.get(host, 0) + 1
            peak_total = max(peak_total, active)
            peak_by_host[host] = max(peak_by_host.get(host, 0), active_by_host[host])
            if active == 8:
                gate.set()
        gate.wait(timeout=2)
        with lock:
            active -= 1
            active_by_host[host] -= 1
        return image_response(content)

    requests = tuple(
        request(f"https://{host}/image-{ordinal}.jpg", sku=f"{host}-{ordinal}")
        for host in ("one.example", "two.example")
        for ordinal in range(1, 5)
    )
    result = mocked_download(requests, handler)

    assert len(result.images) == 8
    assert peak_total == 8
    assert peak_by_host == {"one.example": 4, "two.example": 4}
