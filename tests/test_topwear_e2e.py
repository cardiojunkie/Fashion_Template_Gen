from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

from fashion_cms.catalog_service import (
    TOPWEAR_HEADERS,
    build_qc_report,
    build_topwear_workbook,
    fake_catalog_client,
    generate_catalog_batch,
)
from fashion_cms.database import JobDatabase
from fashion_cms.excel_service import parse_input_workbook
from fashion_cms.image_service import parse_uploaded_images
from fashion_cms.jobs import JobService
from fashion_cms.llm_service import FakeLLMClient, LLMResponse
from fashion_cms.models import AnalysisMode
from fashion_cms.registry import load_registry
from fashion_cms.review import (
    ReviewAction,
    accepted_facts,
    load_review_items,
    persist_review_decision,
    unresolved_review_items,
)
from fashion_cms.topwear_extraction import (
    PROMPT_VERSION,
    SCHEMA_VERSION,
    TOPWEAR_PROFILE_ID,
    run_topwear_job,
)
from fashion_cms.variant_service import build_variant_groups


ROOT = Path(__file__).parents[1]


def workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "sku",
            "base_code",
            "attributes__lulu_ean",
            "attributes__shipping_weight",
            "model_code_input_data",
        ]
    )
    common = {
        "brand": "Acme",
        "product_type": "T-Shirt",
        "material": "Cotton",
        "model": "M1",
        "fit_type": "A-Line Fit",
    }
    worksheet.append(
        ["0001-S", "BASE", "00000001", 0, json.dumps({**common, "size": "S"})]
    )
    worksheet.append(
        ["0001-M", "BASE", "00000002", "1.25", json.dumps({**common, "size": "M"})]
    )
    for cell in (*worksheet[2][:3], *worksheet[3][:3]):
        cell.number_format = "@"
        cell.data_type = "s"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def image_bytes() -> bytes:
    output = BytesIO()
    Image.new("RGB", (20, 20), "blue").save(output, format="JPEG")
    return output.getvalue()


def shared_visual_client() -> FakeLLMClient:
    def responder(request) -> LLMResponse:
        contract = request.contract

        def observation(header: str, value: str) -> dict[str, object]:
            return {
                "header": header,
                "raw_value": value,
                "canonical_value": None,
                "status": "observed",
                "evidence_type": "image",
                "evidence_refs": [contract.image_ids[0]],
                "confidence": "high",
                "normalization_rule": None,
                "note": None,
            }

        output = {
            "attribute_set_id": "topwear",
            "product_profile": contract.product_profile,
            "analysis_mode": contract.analysis_mode.value,
            "group_key": contract.group_key,
            "representative_sku": contract.representative_sku,
            "image_ids": list(contract.image_ids),
            "shared_attributes": [
                observation("attributes__color", "Blue"),
                observation("attributes__design", "Chest Graphic"),
                observation("attributes__neckline", "Crew Neck"),
            ],
            "sku_attributes": [
                {"sku": sku, "observations": []} for sku in contract.represented_skus
            ],
            "warnings": [],
            "conflicts": [],
        }
        return LLMResponse(
            request_id="e2e-vision",
            model="phase5-fake",
            status="completed",
            output_text=json.dumps(output),
        )

    return FakeLLMClient(responder=responder)


def test_upload_to_review_copy_export_and_qc_end_to_end(tmp_path: Path) -> None:
    registry = load_registry(ROOT / "config" / "attribute_registry.xlsx")
    parsed = parse_input_workbook(workbook_bytes(), "input.xlsx")
    images = parse_uploaded_images(
        (("0001-S-1.jpg", image_bytes()),), tuple(row.sku for row in parsed.rows)
    )
    assert parsed.ready and images.ready
    assert [row.sku for row in parsed.rows] == ["0001-S", "0001-M"]
    assert images.images[0].sku == "0001-S"

    database_path = tmp_path / "e2e.sqlite3"
    database = JobDatabase(database_path)
    job_id = JobService(database).create_job(
        parsed.rows,
        images.images,
        attribute_set="topwear",
        registry_version=registry.fingerprint,
        product_profile=TOPWEAR_PROFILE_ID,
        prompt_version=PROMPT_VERSION,
        schema_version=SCHEMA_VERSION,
        model_identifier="phase5-fake",
        image_detail="high",
        modes={"base:BASE": AnalysisMode.BASE_CODE_SIZE_ONLY},
    )
    vision = shared_visual_client()
    run_topwear_job(database, job_id, vision, images.images, registry)
    assert len(vision.calls) == 1

    review_items = load_review_items(database, job_id, registry)
    inherited_colors = [
        item for item in review_items if item.header == "attributes__color"
    ]
    assert {item.sku for item in inherited_colors} == {"0001-S", "0001-M"}
    assert all(item.image_inferred_color for item in inherited_colors)
    assert all(item.evidence_references == ("0001-S-1",) for item in inherited_colors)
    assert all(item.requires_review for item in inherited_colors)

    for item in review_items:
        if item.requires_review:
            persist_review_decision(
                database,
                item,
                ReviewAction.ACCEPT if item.proposed_value else ReviewAction.BLANK,
                registry,
            )
    review_items = load_review_items(database, job_id, registry)
    assert unresolved_review_items(review_items) == ()
    facts = accepted_facts(review_items)
    assert facts["0001-S"]["attributes__size"] == "S"
    assert facts["0001-M"]["attributes__size"] == "M"
    assert facts["0001-S"]["attributes__fit_type"] == "A-Line"

    copy_client = fake_catalog_client()
    catalogs = generate_catalog_batch(
        parsed.rows,
        facts,
        registry,
        copy_client,
        model="phase6-fake",
        groups=database.load_groups(job_id),
    )
    assert len(copy_client.calls) == 1
    assert catalogs["0001-S"].content == catalogs["0001-M"].content
    assert not {"S", "M"} & set(catalogs["0001-S"].content.keywords.split(", "))
    assert ", S," in catalogs["0001-S"].title
    assert ", M," in catalogs["0001-M"].title

    cms = build_topwear_workbook(parsed.rows, review_items, catalogs, registry)
    qc = build_qc_report(review_items)
    reopened = load_workbook(BytesIO(cms), data_only=False)
    worksheet = reopened["CMS Upload"]
    headers = tuple(cell.value for cell in worksheet[1])
    assert headers == TOPWEAR_HEADERS
    assert worksheet.max_row == 3
    assert [worksheet.cell(row, 1).value for row in (2, 3)] == ["0001-S", "0001-M"]
    assert [worksheet.cell(row, 3).value for row in (2, 3)] == ["00000001", "00000002"]
    color_column = headers.index("attributes__color") + 1
    size_column = headers.index("attributes__size") + 1
    assert [worksheet.cell(row, size_column).value for row in (2, 3)] == ["S", "M"]
    assert all(
        str(worksheet.cell(row, color_column).fill.fgColor.rgb).endswith("FFFF00")
        for row in (2, 3)
    )
    reopened.close()

    qc_workbook = load_workbook(BytesIO(qc), data_only=False)
    qc_sheet = qc_workbook["Topwear QC"]
    qc_headers = [cell.value for cell in qc_sheet[1]]
    note_column = qc_headers.index("Image-inference note") + 1
    notes = [
        qc_sheet.cell(row, note_column).value
        for row in range(2, qc_sheet.max_row + 1)
    ]
    assert sum(bool(note) for note in notes) == 2
    assert all("Blue" in note for note in notes if note)
    qc_workbook.close()


def test_color_or_design_differences_prevent_catalog_reuse() -> None:
    registry = load_registry(ROOT / "config" / "attribute_registry.xlsx")
    parsed = parse_input_workbook(workbook_bytes(), "input.xlsx")
    facts = {
        "0001-S": {
            "attributes__brand": "Acme",
            "attributes__product_type": "T-Shirt",
            "attributes__size": "S",
            "attributes__color": "Blue",
            "attributes__design": "Graphic",
        },
        "0001-M": {
            "attributes__brand": "Acme",
            "attributes__product_type": "T-Shirt",
            "attributes__size": "M",
            "attributes__color": "Red",
            "attributes__design": "Plain",
        },
    }
    client = fake_catalog_client()
    groups = build_variant_groups(
        parsed.rows,
        modes={"base:BASE": AnalysisMode.BASE_CODE_SIZE_ONLY},
    )

    generate_catalog_batch(
        parsed.rows,
        facts,
        registry,
        client,
        model="phase6-fake",
        groups=groups,
    )

    assert len(client.calls) == 2
