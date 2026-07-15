import re
import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fashion_cms.registry import (
    DataType,
    EvidencePolicy,
    RegistryValidationError,
    applicable_profile_headers,
    configuration_issues,
    load_registry,
    normalize_value,
    profile_ids,
)


ROOT = Path(__file__).parents[1]
REGISTRY_PATH = ROOT / "config" / "attribute_registry.xlsx"
EXPECTED_HEADER_COUNTS = {
    "topwear": 45,
    "bottomwear": 43,
    "ethnic_wear": 44,
    "inner_sleepwear": 43,
    "footwear": 46,
    "sports_activewear": 46,
    "mens_accessories": 61,
}
ACCESSORY_COMMON_HEADERS = {
    "attributes__brand",
    "attributes__product_type",
    "attributes__model",
    "attributes__color",
    "attributes__size",
    "attributes__gender",
    "attributes__age_group",
    "attributes__season",
    "attributes__occasion",
    "attributes__occasion_type",
    "attributes__package_contents",
    "attributes__in_the_box",
    "attributes__country_of_origin",
    "attributes__weight",
    "attributes__product_dimensions",
    "attributes__other_information",
}
ACCESSORY_PROFILE_HEADERS = {
    "bags_luggage": ACCESSORY_COMMON_HEADERS
    | {
        "attributes__material",
        "attributes__outer_material",
        "attributes__inner_material",
        "attributes__care_instructions",
        "attributes__bag_type",
        "attributes__closure",
        "attributes__fastening_type",
        "attributes__lock_type",
        "attributes__tsa_combination_lock",
        "attributes__strap_type",
        "attributes__compartments",
        "attributes__laptop_compartment",
        "attributes__no_of_pockets",
        "attributes__pattern",
        "attributes__pattern_type",
        "attributes__design",
        "attributes__water_resistance",
    },
    "caps_headwear": ACCESSORY_COMMON_HEADERS
    | {
        "attributes__material",
        "attributes__fabric",
        "attributes__fabric_care",
        "attributes__care_instructions",
        "attributes__cap_type",
        "attributes__closure",
        "attributes__fastening_type",
        "attributes__pattern",
        "attributes__pattern_type",
        "attributes__design",
        "attributes__water_resistance",
    },
    "watches": ACCESSORY_COMMON_HEADERS
    | {
        "attributes__material",
        "attributes__closure",
        "attributes__fastening_type",
        "attributes__strap_type",
        "attributes__case_size",
        "attributes__band_size",
        "attributes__movement_type",
        "attributes__display_feature",
        "attributes__water_resistance",
    },
    "eyewear": ACCESSORY_COMMON_HEADERS
    | {
        "attributes__material",
        "attributes__polarization",
        "attributes__lens_color",
        "attributes__lens_shape",
        "attributes__lens_type",
        "attributes__frame_color",
        "attributes__frame_material",
        "attributes__frame_shape",
        "attributes__frame_size",
    },
    "belts_wallets_ties_other": ACCESSORY_COMMON_HEADERS
    | {
        "attributes__material",
        "attributes__fabric",
        "attributes__fabric_care",
        "attributes__care_instructions",
        "attributes__closure",
        "attributes__fastening_type",
        "attributes__pattern",
        "attributes__pattern_type",
        "attributes__design",
    },
}


def expected_attribute_sets() -> dict[str, tuple[str, tuple[str, ...]]]:
    plan = (ROOT / "PLAN.md").read_text()
    blocks = re.findall(
        r"## A\d+\. (.*?) \(`([^`]+)`\)\n\n```text\n(.*?)\n```",
        plan,
        re.DOTALL,
    )
    return {set_id: (name, tuple(headers.splitlines())) for name, set_id, headers in blocks}


def workbook_fixture(tmp_path: Path, mutate) -> Path:
    path = tmp_path / "registry.xlsx"
    shutil.copy2(REGISTRY_PATH, path)
    workbook = load_workbook(path)
    mutate(workbook)
    workbook.save(path)
    return path


def row_with_value(worksheet, column: int, value: str) -> int:
    return next(
        row for row in range(2, worksheet.max_row + 1) if worksheet.cell(row, column).value == value
    )


def test_committed_header_names_membership_and_order_match_plan() -> None:
    registry = load_registry(REGISTRY_PATH)
    expected = expected_attribute_sets()

    assert list(registry.mappings_by_set) == list(expected)
    for set_id, (name, headers) in expected.items():
        rows = [row for row in registry.attribute_sets if row.attribute_set_id == set_id]
        assert {row.attribute_set_name for row in rows} == {name}
        assert len(headers) == EXPECTED_HEADER_COUNTS[set_id]
        assert registry.mappings_by_set[set_id] == headers


def test_committed_registry_has_one_correct_definition_per_header() -> None:
    registry = load_registry(REGISTRY_PATH)
    mapped_headers = {header for headers in registry.mappings_by_set.values() for header in headers}

    assert len(registry.definitions) == len(mapped_headers) == 78
    assert set(registry.definitions_by_header) == mapped_headers
    for header in {
        "sku",
        "base_code",
        "attributes__lulu_ean",
        "attributes__shipping_weight",
    }:
        definition = registry.definitions_by_header[header]
        assert definition.data_type == DataType.SYSTEM_COPY
        assert definition.evidence_policy == EvidencePolicy.SYSTEM_COPY
    for header in {
        "attributes__keywords",
        "name",
        "attributes__product_title",
        *(f"attributes__bullet_point_{number}" for number in range(1, 7)),
    }:
        definition = registry.definitions_by_header[header]
        assert definition.data_type == DataType.GENERATED_TEXT
        assert definition.evidence_policy == EvidencePolicy.GENERATED_CONTENT

    assert registry.permitted_values_by_header["attributes__color"] == (
        "Blue",
        "Red",
        "White",
        "Black",
        "Green",
        "Grey",
        "Brown",
    )
    assert registry.definitions_by_header["attributes__color"].data_type == DataType.ENUM
    assert registry.definitions_by_header["attributes__fit_type"].data_type == DataType.ENUM
    assert registry.permitted_values_by_header["attributes__fit_type"] == ("A-Line",)
    assert {row.attribute_header for row in registry.permitted_values if row.values} == {
        "attributes__color",
        "attributes__fit_type",
    }
    assert len(registry.aliases) == 1
    assert registry.aliases[0].active is True
    assert registry.aliases_by_header == {"attributes__fit_type": {"a line fit": "A-Line"}}
    assert len(registry.fingerprint) == 64


def test_committed_technical_profiles_and_configuration_health() -> None:
    registry = load_registry(REGISTRY_PATH)

    assert profile_ids(registry, "topwear") == ("topwear_mvp",)
    assert configuration_issues(registry, "topwear") == ()
    for set_id in (
        "bottomwear",
        "ethnic_wear",
        "inner_sleepwear",
        "footwear",
        "sports_activewear",
    ):
        assert profile_ids(registry, set_id) == (f"{set_id}_default",)
        expected = tuple(
            header
            for header in registry.mappings_by_set[set_id]
            if registry.definitions_by_header[header].data_type
            not in {DataType.SYSTEM_COPY, DataType.GENERATED_TEXT}
        )
        assert applicable_profile_headers(registry, set_id, f"{set_id}_default") == expected
        assert configuration_issues(registry, set_id) == (
            "Approved CMS product types are absent.",
            "Approved set-specific permitted-value sources are absent.",
        )

    assert profile_ids(registry, "mens_accessories") == tuple(ACCESSORY_PROFILE_HEADERS)
    assert configuration_issues(registry, "mens_accessories") == (
        "Approved CMS product types are absent.",
        "Approved set-specific permitted-value sources are absent.",
    )
    assert {
        profile.product_type
        for profile in registry.profiles
        if profile.attribute_set_id != "topwear"
    } == {None}


def test_accessory_profiles_have_explicit_isolated_applicability() -> None:
    registry = load_registry(REGISTRY_PATH)
    eligible = {
        header
        for header in registry.mappings_by_set["mens_accessories"]
        if registry.definitions_by_header[header].data_type
        not in {DataType.SYSTEM_COPY, DataType.GENERATED_TEXT}
    }

    for profile_id, expected in ACCESSORY_PROFILE_HEADERS.items():
        rows = registry.profiles_by_id[("mens_accessories", profile_id)]
        assert {row.header for row in rows} == eligible
        assert set(applicable_profile_headers(registry, "mens_accessories", profile_id)) == expected


def test_profile_helpers_reject_unknown_references() -> None:
    registry = load_registry(REGISTRY_PATH)

    with pytest.raises(ValueError, match="Unknown attribute set"):
        profile_ids(registry, "missing")
    with pytest.raises(ValueError, match="Unknown profile"):
        applicable_profile_headers(registry, "bottomwear", "missing")
    with pytest.raises(ValueError, match="Unknown attribute set"):
        configuration_issues(registry, "missing")


def test_normalization_handles_unicode_case_whitespace_and_punctuation() -> None:
    assert normalize_value("  Ａ-Line__FIT  ") == normalize_value("a line fit")


def test_duplicate_header_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Attribute_Sets"]
        sheet.append(["topwear", "Topwear", 46, "attributes__color", False])

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="duplicate headers"):
        load_registry(path)


def test_duplicate_canonical_value_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        definitions = workbook["Attribute_Definitions"]
        definition_row = row_with_value(definitions, 1, "attributes__fit_type")
        definitions.cell(definition_row, 2).value = "ENUM"

        values = workbook["Permitted_Values"]
        values.cell(1, 4).value = "value_2"
        value_row = row_with_value(values, 1, "attributes__fit_type")
        values.cell(value_row, 2).value = "ENUM"
        values.cell(value_row, 3).value = "A-Line"
        values.cell(value_row, 4).value = "a line"

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="duplicate canonical values"):
        load_registry(path)


def test_missing_definition_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Attribute_Definitions"]
        sheet.delete_rows(row_with_value(sheet, 1, "attributes__color"))

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="missing definitions"):
        load_registry(path)


def test_active_alias_without_canonical_value_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        workbook["Value_Aliases"].cell(2, 3).value = "Missing"

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="points to missing canonical value"):
        load_registry(path)


def test_inactive_alias_without_canonical_value_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        workbook["Value_Aliases"].cell(2, 3).value = "Missing"
        workbook["Value_Aliases"].cell(2, 4).value = False

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="points to missing canonical value"):
        load_registry(path)


def test_invalid_data_type_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Attribute_Definitions"]
        row = row_with_value(sheet, 1, "attributes__color")
        sheet.cell(row, 2).value = "NOT_A_DATA_TYPE"

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="Input should be"):
        load_registry(path)


def test_invalid_profile_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        workbook["Product_Profiles"].append(
            ["topwear", "shirt", "shirts", "attributes__not_a_header", True]
        )

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="is not in topwear"):
        load_registry(path)


def test_duplicate_profile_mapping_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Product_Profiles"]
        sheet.append([cell.value for cell in sheet[2]])

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="duplicate mapping"):
        load_registry(path)


def test_inconsistent_profile_product_type_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Product_Profiles"]
        row = next(
            number
            for number in range(2, sheet.max_row + 1)
            if sheet.cell(number, 1).value == "bottomwear"
        )
        sheet.cell(row, 2).value = "Unapproved Product Type"

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="inconsistent product types"):
        load_registry(path)


def test_incomplete_profile_matrix_is_rejected(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Product_Profiles"]
        row = next(
            number
            for number in range(2, sheet.max_row + 1)
            if sheet.cell(number, 1).value == "bottomwear"
            and sheet.cell(number, 4).value == "attributes__brand"
        )
        sheet.delete_rows(row)

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="is missing headers"):
        load_registry(path)


def test_mandatory_accessory_profile_is_required(tmp_path: Path) -> None:
    def mutate(workbook) -> None:
        sheet = workbook["Product_Profiles"]
        for row in range(sheet.max_row, 1, -1):
            if (
                sheet.cell(row, 1).value == "mens_accessories"
                and sheet.cell(row, 3).value == "eyewear"
            ):
                sheet.delete_rows(row)

    path = workbook_fixture(tmp_path, mutate)
    with pytest.raises(RegistryValidationError, match="missing mandatory profiles.*eyewear"):
        load_registry(path)
