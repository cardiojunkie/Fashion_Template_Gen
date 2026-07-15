from pathlib import Path

import pytest

from fashion_cms.normalization import MatchMethod, normalize_attribute_value
from fashion_cms.registry import RegistryValidationError, load_registry


ROOT = Path(__file__).parents[1]


@pytest.fixture(scope="module")
def registry():
    return load_registry(ROOT / "config" / "attribute_registry.xlsx")


def test_exact_and_normalized_canonical_matches(registry) -> None:
    exact = normalize_attribute_value(registry, "attributes__color", "Blue")
    case = normalize_attribute_value(registry, "attributes__color", "blue")
    whitespace = normalize_attribute_value(registry, "attributes__color", "  Blue  ")
    punctuation = normalize_attribute_value(registry, "attributes__fit_type", "A Line")

    assert (exact.canonical_value, exact.method) == ("Blue", MatchMethod.EXACT_CANONICAL)
    assert (case.canonical_value, case.method) == (
        "Blue",
        MatchMethod.NORMALIZED_CANONICAL,
    )
    assert whitespace.canonical_value == "Blue"
    assert punctuation.canonical_value == "A-Line"


def test_approved_alias_is_header_scoped(registry) -> None:
    mapped = normalize_attribute_value(
        registry, "attributes__fit_type", "A-Line Fit"
    )
    wrong_header = normalize_attribute_value(
        registry, "attributes__color", "A-Line Fit"
    )

    assert mapped.canonical_value == "A-Line"
    assert mapped.method == MatchMethod.APPROVED_ALIAS
    assert mapped.alias_used == "A-Line Fit"
    assert wrong_header.canonical_value is None


def test_unknown_value_is_unmapped_and_fuzzy_is_review_only(registry) -> None:
    unknown = normalize_attribute_value(registry, "attributes__color", "Chartreuse")
    fuzzy = normalize_attribute_value(registry, "attributes__color", "Blaack")

    assert unknown.canonical_value is None
    assert fuzzy.canonical_value is None
    assert fuzzy.fuzzy_suggestion == "Black"
    assert fuzzy.method == MatchMethod.FUZZY_SUGGESTION


def test_ambiguous_fuzzy_match_has_no_suggestion(registry) -> None:
    values = dict(registry.permitted_values_by_header)
    values["attributes__color"] = ("Blue One", "Blue Two")
    ambiguous_registry = registry.model_copy(
        update={"permitted_values_by_header": values}
    )

    result = normalize_attribute_value(
        ambiguous_registry,
        "attributes__color",
        "Blue",
        fuzzy_threshold=0.5,
    )

    assert result.ambiguous is True
    assert result.canonical_value is None
    assert result.fuzzy_suggestion is None


def test_registry_rejects_normalized_duplicate_permitted_values(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    source = ROOT / "config" / "attribute_registry.xlsx"
    destination = tmp_path / "registry.xlsx"
    destination.write_bytes(source.read_bytes())
    workbook = load_workbook(destination)
    values = workbook["Permitted_Values"]
    values.cell(1, 4).value = "value_2"
    row = next(
        number
        for number in range(2, values.max_row + 1)
        if values.cell(number, 1).value == "attributes__fit_type"
    )
    values.cell(row, 4).value = "a line"
    workbook.save(destination)
    workbook.close()

    with pytest.raises(RegistryValidationError, match="duplicate canonical values"):
        load_registry(destination)
