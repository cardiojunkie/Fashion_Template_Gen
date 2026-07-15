import re
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from fashion_cms.excel_service import (
    REQUIRED_COLUMNS,
    SYSTEM_COPY_FIELDS,
    build_blank_cms_workbook,
    parse_input_workbook,
)
from fashion_cms.models import InputRow
from fashion_cms.registry import load_registry


ROOT = Path(__file__).parents[1]


def workbook_bytes(rows: list[list[object]], headers: tuple[str, ...] = REQUIRED_COLUMNS) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def issue_codes(result) -> set[str]:
    return {issue.code for issue in result.issues}


def rewrite_worksheet_xml(content: bytes, mutate) -> bytes:
    source = ZipFile(BytesIO(content))
    output = BytesIO()
    with source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            data = source.read(member)
            if member.filename == "xl/worksheets/sheet1.xml":
                data = mutate(data)
            target.writestr(member, data)
    return output.getvalue()


def test_leading_zero_identifiers_survive_parse_and_exact_export() -> None:
    result = parse_input_workbook(
        workbook_bytes([["0007", "001", "0000123", "0.25", "shirt"]]),
        "input.xlsx",
    )
    assert result.ready
    assert result.rows[0].sku == "0007"
    assert result.rows[0].base_code == "001"
    assert result.rows[0].attributes__lulu_ean == "0000123"

    headers = load_registry(ROOT / "config" / "attribute_registry.xlsx").mappings_by_set["topwear"]
    output = build_blank_cms_workbook(result.rows, headers)
    workbook = load_workbook(BytesIO(output), data_only=False)
    worksheet = workbook.active
    assert workbook.sheetnames == ["CMS Upload"]
    assert tuple(cell.value for cell in worksheet[1]) == headers
    assert worksheet.max_column == len(headers)
    assert worksheet.max_row == 2

    output_values = {
        header: worksheet.cell(2, column + 1).value for column, header in enumerate(headers)
    }
    assert output_values["sku"] == "0007"
    assert output_values["base_code"] == "001"
    assert output_values["attributes__lulu_ean"] == "0000123"
    assert output_values["attributes__shipping_weight"] == "0.25"
    assert all(
        output_values[header] is None for header in headers if header not in SYSTEM_COPY_FIELDS
    )
    for header in ("sku", "base_code", "attributes__lulu_ean"):
        cell = worksheet.cell(2, headers.index(header) + 1)
        assert cell.data_type == "s"
        assert cell.number_format == "@"
    workbook.close()


def test_blank_base_code_and_duplicate_ean_warn_without_blocking() -> None:
    result = parse_input_workbook(
        workbook_bytes(
            [
                ["SKU-1", None, "0001", 1.2, "shirt"],
                ["SKU-2", "BASE", "0001", 1.3, "shirt"],
            ]
        ),
        "input.xlsx",
    )
    assert result.ready
    assert result.rows[0].group_key == "SKU-1"
    assert {"BLANK_BASE_CODE", "DUPLICATE_EAN"} <= issue_codes(result)


def test_formulas_numeric_identifiers_and_duplicate_skus_are_critical() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(REQUIRED_COLUMNS)
    worksheet.append(["DUP", "BASE", "0001", 1, "shirt"])
    worksheet.append(["DUP", "BASE", "0002", 1, "shirt"])
    worksheet.append([123, "BASE", "0003", 1, "shirt"])
    worksheet.append(["FORMULA", "=A2", "0004", 1, "shirt"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    result = parse_input_workbook(output.getvalue(), "input.xlsx")
    assert not result.ready
    assert {"DUPLICATE_SKU", "INVALID_CELL", "FORMULA_NOT_ALLOWED"} <= issue_codes(result)


@pytest.mark.parametrize(
    ("content", "filename", "expected_code"),
    [
        (b"not an xlsx", "input.xlsx", "MALFORMED_WORKBOOK"),
        (workbook_bytes([["SKU"]], ("sku",)), "input.xlsx", "MISSING_COLUMNS"),
        (workbook_bytes([]), "input.xls", "UNSUPPORTED_WORKBOOK_TYPE"),
    ],
)
def test_unsupported_malformed_or_missing_columns_block(
    content: bytes, filename: str, expected_code: str
) -> None:
    result = parse_input_workbook(content, filename)
    assert not result.ready
    assert expected_code in issue_codes(result)


def test_formula_like_identifier_exports_as_literal_text() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(REQUIRED_COLUMNS)
    worksheet.append(["=SKU", "BASE", "0001", 1, "shirt"])
    worksheet["A2"].data_type = "s"
    source = BytesIO()
    workbook.save(source)
    workbook.close()

    result = parse_input_workbook(source.getvalue(), "input.xlsx")
    assert result.ready
    output = build_blank_cms_workbook(result.rows, ("sku",))
    exported = load_workbook(BytesIO(output), data_only=False)
    assert exported.active["A2"].value == "=SKU"
    assert exported.active["A2"].data_type == "s"
    exported.close()


@pytest.mark.parametrize("dimension", [None, b'<dimension ref="A1:E2"/>'])
def test_missing_or_understated_dimensions_do_not_hide_rows(dimension: bytes | None) -> None:
    source = workbook_bytes(
        [
            ["SKU-1", "BASE", "0001", 1, "shirt"],
            ["SKU-2", "BASE", "0002", 1, "shirt"],
        ]
    )

    def mutate(xml: bytes) -> bytes:
        replacement = dimension or b""
        return re.sub(rb'<dimension ref="[^"]+"/>', replacement, xml, count=1)

    result = parse_input_workbook(rewrite_worksheet_xml(source, mutate), "input.xlsx")
    assert result.ready
    assert [row.sku for row in result.rows] == ["SKU-1", "SKU-2"]


def test_lazy_malformed_worksheet_xml_is_actionable_critical_error() -> None:
    source = workbook_bytes([["SKU", "BASE", "0001", 1, "shirt"]])
    malformed = rewrite_worksheet_xml(source, lambda xml: xml[:-20])
    result = parse_input_workbook(malformed, "input.xlsx")
    assert not result.ready
    assert issue_codes(result) == {"MALFORMED_WORKBOOK"}


def test_excel_error_identifier_is_rejected() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(REQUIRED_COLUMNS)
    worksheet.append(["#N/A", "BASE", "0001", 1, "shirt"])
    worksheet["A2"].data_type = "e"
    output = BytesIO()
    workbook.save(output)
    workbook.close()

    result = parse_input_workbook(output.getvalue(), "input.xlsx")
    assert not result.ready
    assert "INVALID_CELL" in issue_codes(result)


def test_export_rejects_values_that_excel_would_silently_truncate() -> None:
    row = InputRow.model_construct(
        row_number=2,
        sku="S" * 32_768,
        base_code=None,
        attributes__lulu_ean=None,
        attributes__shipping_weight=None,
        model_code_input_data=None,
    )
    with pytest.raises(ValueError, match="character limit"):
        build_blank_cms_workbook((row,), ("sku",))
