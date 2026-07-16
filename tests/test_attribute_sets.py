from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fashion_cms.catalog_service import (
    build_cms_workbook,
    fake_catalog_client,
    generate_catalog_batch,
)
from fashion_cms.database import JobDatabase
from fashion_cms.jobs import JobService
from fashion_cms.llm_service import LLMResponse
from fashion_cms.models import AnalysisMode, InputRow, UploadedImage
from fashion_cms.registry import applicable_profile_headers, load_registry
from fashion_cms.review import (
    ReviewAction,
    accepted_facts,
    load_review_items,
    persist_review_decision,
    unresolved_review_items,
)
from fashion_cms.topwear_extraction import (
    ATTRIBUTE_PROMPT_VERSION,
    ATTRIBUTE_SCHEMA_VERSION,
    AttributeResultError,
    EvidenceType,
    ExtractionRecord,
    ObservationStatus,
    applicable_attribute_headers,
    build_attribute_request,
    fake_attribute_client,
    run_attribute_job,
    validate_attribute_response,
)
from fashion_cms.variant_service import CacheContext, build_request_plan, build_variant_groups


ROOT = Path(__file__).parents[1]
REGISTRY_PATH = ROOT / "config" / "attribute_registry.xlsx"
GOLDEN = json.loads((ROOT / "tests/fixtures/phase7_golden.json").read_text())
CASES = tuple(
    (case["attribute_set"], case["profile"], case["product_type"])
    for case in GOLDEN["cases"]
)


@pytest.fixture(scope="module")
def registry():
    return load_registry(REGISTRY_PATH)


def row(
    sku: str,
    base_code: str,
    data: dict[str, str],
    row_number: int,
) -> InputRow:
    return InputRow(
        row_number=row_number,
        sku=sku,
        base_code=base_code,
        attributes__lulu_ean=f"00{row_number}",
        attributes__shipping_weight="1.0",
        input_data=json.dumps(data),
    )


def image(sku: str) -> UploadedImage:
    return UploadedImage(
        source_name=f"{sku}-1.jpg",
        filename=f"{sku}-1.jpg",
        sku=sku,
        ordinal=1,
        image_format="jpeg",
        width=10,
        height=10,
        content=f"image:{sku}".encode(),
    )


def context(registry, attribute_set: str, profile: str) -> CacheContext:
    return CacheContext(
        attribute_set=attribute_set,
        product_profile=profile,
        registry_version=registry.fingerprint,
        prompt_version=ATTRIBUTE_PROMPT_VERSION,
        schema_version=ATTRIBUTE_SCHEMA_VERSION,
        model_identifier="phase7-fake",
        image_detail="high",
    )


def request_for(registry, attribute_set: str, profile: str, data=None):
    product_row = row("SKU-A", "BASE", data or {}, 2)
    product_image = image(product_row.sku)
    request_context = context(registry, attribute_set, profile)
    item = build_request_plan(
        build_variant_groups((product_row,), (product_image,), product_profile=profile),
        request_context,
    ).items[0]
    request = build_attribute_request(
        item,
        (product_row,),
        (product_image,),
        registry,
        request_context,
    )
    return request, request_context


def wire(request, *, shared=(), by_sku=None, **changes):
    contract = request.contract
    result = {
        "attribute_set_id": contract.attribute_set_id,
        "product_profile": contract.product_profile,
        "analysis_mode": contract.analysis_mode.value,
        "group_key": contract.group_key,
        "representative_sku": contract.representative_sku,
        "image_ids": list(contract.image_ids),
        "shared_attributes": list(shared),
        "sku_attributes": [
            {"sku": sku, "observations": list((by_sku or {}).get(sku, ()))}
            for sku in contract.represented_skus
        ],
        "warnings": [],
        "conflicts": [],
    }
    result.update(changes)
    return result


def observation(header: str, value: str, image_id: str) -> dict[str, object]:
    return {
        "header": header,
        "raw_value": value,
        "canonical_value": None,
        "status": "observed",
        "evidence_type": "image",
        "evidence_refs": [image_id],
        "confidence": "high",
        "normalization_rule": None,
        "note": None,
    }


def response(payload: object) -> LLMResponse:
    return LLMResponse(
        request_id="phase7-test",
        model="phase7-fake",
        status="completed",
        output_text=payload if isinstance(payload, str) else json.dumps(payload),
    )


@pytest.mark.parametrize(("attribute_set", "profile", "product_type"), CASES)
def test_each_set_runs_size_only_review_copy_and_exact_export(
    tmp_path: Path,
    registry,
    attribute_set: str,
    profile: str,
    product_type: str,
) -> None:
    common = {
        "brand": "Acme",
        "product_type": product_type,
        "color": GOLDEN["size_only"]["color"],
        "material": "Cotton",
    }
    rows = (
        row(f"{attribute_set}-N", "NORMAL", {**common, "size": "L"}, 2),
        row(f"{attribute_set}-S", "SIZE", {**common, "size": "S"}, 3),
        row(f"{attribute_set}-M", "SIZE", {**common, "size": "M"}, 4),
    )
    images = tuple(image(product_row.sku) for product_row in rows)
    database = JobDatabase(tmp_path / f"{attribute_set}.sqlite3")
    job_id = JobService(database).create_job(
        rows,
        images,
        attribute_set=attribute_set,
        product_profile=profile,
        registry_version=registry.fingerprint,
        prompt_version=ATTRIBUTE_PROMPT_VERSION,
        schema_version=ATTRIBUTE_SCHEMA_VERSION,
        model_identifier="phase7-fake",
        modes={"base:SIZE": AnalysisMode.BASE_CODE_SIZE_ONLY},
    )

    client = fake_attribute_client()
    completed = run_attribute_job(database, job_id, client, images, registry)
    assert completed.status.value == "REVIEW_REQUIRED"
    assert len(client.calls) == 2
    assert any(call.contract.analysis_mode == AnalysisMode.BASE_CODE_SIZE_ONLY for call in client.calls)

    items = load_review_items(database, job_id, registry)
    for item in items:
        persist_review_decision(
            database,
            item,
            ReviewAction.ACCEPT if item.proposed_value else ReviewAction.BLANK,
            registry,
        )
    items = load_review_items(database, job_id, registry)
    assert unresolved_review_items(items) == ()
    facts = accepted_facts(items)
    catalogs = generate_catalog_batch(
        rows,
        facts,
        registry,
        fake_catalog_client(),
        model="phase7-fake",
        groups=database.load_groups(job_id),
        attribute_set=attribute_set,
        product_profile=profile,
    )
    exported = build_cms_workbook(
        rows,
        items,
        catalogs,
        registry,
        attribute_set=attribute_set,
        product_profile=profile,
    )
    workbook = load_workbook(BytesIO(exported), data_only=False)
    worksheet = workbook["CMS Upload"]
    assert tuple(cell.value for cell in worksheet[1]) == registry.mappings_by_set[attribute_set]
    assert worksheet.max_row == 4
    assert [worksheet.cell(index, 1).value for index in range(2, 5)] == [row.sku for row in rows]
    assert all(worksheet.cell(index, 1).data_type == "s" for index in range(2, 5))
    if attribute_set == "mens_accessories":
        for header in ("attributes__movement_type", "attributes__polarization"):
            column = registry.mappings_by_set[attribute_set].index(header) + 1
            assert all(worksheet.cell(index, column).value is None for index in range(2, 5))
    workbook.close()


@pytest.mark.parametrize(("attribute_set", "profile", "_"), CASES)
def test_each_set_rejects_unknown_and_malformed_model_output(
    registry, attribute_set: str, profile: str, _: str
) -> None:
    request, request_context = request_for(registry, attribute_set, profile)
    unknown = observation("attributes__not_real", "value", request.contract.image_ids[0])
    with pytest.raises(AttributeResultError):
        validate_attribute_response(
            response(wire(request, by_sku={"SKU-A": (unknown,)})),
            request,
            registry,
            job_id="job",
            context=request_context,
        )
    with pytest.raises(AttributeResultError):
        validate_attribute_response(
            response("{"), request, registry, job_id="job", context=request_context
        )


@pytest.mark.parametrize(
    ("attribute_set", "profile", "header"),
    (
        ("footwear", "footwear_default", "attributes__heel_height"),
        ("footwear", "footwear_default", "attributes__outer_material"),
        ("sports_activewear", "sports_activewear_default", "attributes__elasticity"),
        ("mens_accessories", "bags_luggage", "attributes__tsa_combination_lock"),
        ("mens_accessories", "watches", "attributes__movement_type"),
        ("mens_accessories", "eyewear", "attributes__polarization"),
        ("mens_accessories", "bags_luggage", "attributes__water_resistance"),
    ),
)
def test_visual_technical_claims_remain_unknown(
    registry, attribute_set: str, profile: str, header: str
) -> None:
    request, request_context = request_for(registry, attribute_set, profile)
    result = validate_attribute_response(
        response(
            wire(
                request,
                by_sku={
                    "SKU-A": (
                        observation(header, "unsupported claim", request.contract.image_ids[0]),
                    )
                },
            )
        ),
        request,
        registry,
        job_id="job",
        context=request_context,
    )
    vision = ExtractionRecord.model_validate(result).vision_result
    found = next(item for item in vision.sku_attributes["SKU-A"] if item.header == header)
    assert found.status == ObservationStatus.UNKNOWN
    assert found.canonical_value is None
    assert found.evidence_type == EvidenceType.IMAGE


@pytest.mark.parametrize(("attribute_set", "profile", "_"), CASES)
def test_explicit_variant_colors_and_profile_conflicts_are_not_shared(
    registry, attribute_set: str, profile: str, _: str
) -> None:
    rows = (
        row("SKU-B", "VARIANT", {"color": "Blue", "product_profile": profile}, 2),
        row("SKU-R", "VARIANT", {"color": "Red", "product_profile": "wrong"}, 3),
    )
    images = (image("SKU-B"), image("SKU-R"))
    groups = build_variant_groups(
        rows,
        images,
        modes={"base:VARIANT": AnalysisMode.BASE_CODE_SIZE_ONLY},
        product_profile=profile,
    )
    assert any(warning.startswith("Multiple colors") for warning in groups[0].warnings)
    assert any(warning.startswith("Product profile conflict") for warning in groups[0].warnings)
    request_context = context(registry, attribute_set, profile)
    item = build_request_plan(groups, request_context).items[0]
    request = build_attribute_request(item, rows, images, registry, request_context)
    visual = observation("attributes__color", "Blue", request.contract.image_ids[0])
    result = validate_attribute_response(
        response(wire(request, shared=(visual,))),
        request,
        registry,
        job_id="job",
        context=request_context,
    )
    vision = ExtractionRecord.model_validate(result).vision_result
    assert all(item.header != "attributes__color" for item in vision.shared_attributes)
    assert {
        sku: next(
            item.canonical_value
            for item in observations
            if item.header == "attributes__color"
        )
        for sku, observations in vision.sku_attributes.items()
    } == {"SKU-B": "Blue", "SKU-R": "Red"}


@pytest.mark.parametrize(
    ("profile", "excluded"),
    (
        ("bags_luggage", ("attributes__movement_type", "attributes__polarization")),
        ("caps_headwear", ("attributes__bag_type", "attributes__case_size")),
        ("watches", ("attributes__bag_type", "attributes__lens_type")),
        ("eyewear", ("attributes__bag_type", "attributes__movement_type")),
        ("belts_wallets_ties_other", ("attributes__bag_type", "attributes__case_size")),
    ),
)
def test_accessory_request_payload_excludes_other_profiles(
    registry, profile: str, excluded: tuple[str, ...]
) -> None:
    request, _ = request_for(registry, "mens_accessories", profile)
    assert "attributes__product_type" not in request.contract.allowed_headers
    assert request.contract.allowed_headers == applicable_attribute_headers(
        registry, "mens_accessories", profile
    )
    assert set(request.contract.allowed_headers) <= set(
        applicable_profile_headers(registry, "mens_accessories", profile)
    )
    payload = json.dumps(request.payload)
    assert all(header not in payload for header in excluded)


@pytest.mark.parametrize(("attribute_set", "profile", "_"), CASES)
def test_missing_approved_product_types_never_enter_extraction_schema(
    registry, attribute_set: str, profile: str, _: str
) -> None:
    request, _ = request_for(registry, attribute_set, profile)
    assert "attributes__product_type" not in request.contract.allowed_headers
    assert "attributes__product_type" not in json.dumps(request.payload)


def test_phase7_golden_fixture_covers_every_new_set_and_variant_mode() -> None:
    assert GOLDEN["fixture_version"] == "1"
    assert {case[0] for case in CASES} == {
        "bottomwear",
        "ethnic_wear",
        "inner_sleepwear",
        "footwear",
        "sports_activewear",
        "mens_accessories",
    }
    assert GOLDEN["size_only"]["sizes"] == ["S", "M"]
    assert len(set(GOLDEN["visually_varying"]["colors"])) > 1
    assert len(set(GOLDEN["visually_varying"]["patterns"])) > 1
