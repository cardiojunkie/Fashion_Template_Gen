from __future__ import annotations

import json
import stat
import threading
import time
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile, ZipInfo

import httpx
import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image
from pydantic import ValidationError

from fashion_cms.catalog_service import (
    build_qc_report,
    fake_catalog_client,
    generate_catalog_content,
)
from fashion_cms.cleanup import cleanup_paths, durable_cleanup_disabled
from fashion_cms.config import (
    ModelPricing,
    ResourceLimits,
    maximum_job_cost,
    usage_cost,
)
from fashion_cms.database import JobDatabase
from fashion_cms.evaluation import (
    ApprovalStatus,
    FieldThreshold,
    ModelEvaluationConfig,
    ThresholdConfiguration,
    ThresholdPolicy,
    compare_models,
    engineering_echo_prediction,
    load_dataset,
    route_threshold_policies,
)
from fashion_cms.excel_service import parse_input_workbook, preflight_xlsx
from fashion_cms.image_downloader import DownloadSettings, download_images
from fashion_cms.image_service import parse_uploaded_images
from fashion_cms.jobs import JobService, fake_extract
from fashion_cms.llm_service import sanitize_error
from fashion_cms.models import DownloadResult, ImageUrlRequest, InputRow, JobStatus, WorkItemStatus
from fashion_cms.registry import load_registry
from fashion_cms.release_gates import GATES, GateStatus, build_report, load_report


ROOT = Path(__file__).parents[1]
DATASET = ROOT / "tests" / "fixtures" / "phase8_engineering_dataset.json"
PUBLIC_URL = "https://images.example/item.jpg"


def _model(model_id: str) -> ModelEvaluationConfig:
    return ModelEvaluationConfig(
        model_id=model_id,
        configuration={"temperature": 0},
        prompt_version="phase8-fake-prompt-v1",
        schema_version="phase8-fake-schema-v1",
        registry_version="phase8-registry",
        image_detail="high",
    )


def _image_bytes(
    image_format: str = "PNG",
    *,
    size: tuple[int, int] = (12, 8),
    mode: str = "RGB",
) -> bytes:
    image = Image.new(mode, size, "blue" if mode == "RGB" else 1)
    output = BytesIO()
    image.save(output, format=image_format)
    image.close()
    return output.getvalue()


def _zip(*members: tuple[str | ZipInfo, bytes]) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w") as archive:
        for name, content in members:
            archive.writestr(name, content)
    return output.getvalue()


def _rows(count: int) -> tuple[InputRow, ...]:
    return tuple(
        InputRow(row_number=index + 2, sku=f"SKU-{index}", base_code=f"BASE-{index}")
        for index in range(count)
    )


def test_engineering_dataset_and_fake_comparison_cover_release_dimensions() -> None:
    dataset = load_dataset(DATASET)
    report = compare_models(
        dataset,
        (_model("deterministic-a"), _model("deterministic-b")),
        engineering_echo_prediction,
    )

    assert dataset.approval_status == ApprovalStatus.PENDING
    assert {case.attribute_set for case in dataset.cases} == {
        "topwear",
        "bottomwear",
        "ethnic_wear",
        "inner_sleepwear",
        "footwear",
        "sports_activewear",
        "mens_accessories",
    }
    assert {
        case.product_profile
        for case in dataset.cases
        if case.attribute_set == "mens_accessories"
    } == {
        "bags_luggage",
        "caps_headwear",
        "watches",
        "eyewear",
        "belts_wallets_ties_other",
    }
    tags = {tag for case in dataset.cases for tag in case.scenario_tags}
    assert {
        "per_sku",
        "valid_size_only",
        "visually_different_variant",
        "missing_images",
        "conflicting_input_visual",
        "invalid_unmapped_value",
        "conservative_explicit_only",
        "partial_failure",
        "prompt_injection",
    } <= tags
    assert len(report.runs) == 2
    assert report.overall.metrics["precision"].value == 1
    assert report.overall.metrics["precision"].sample_count > 0
    assert report.overall.metrics["variant_leakage_rate"].value == 0
    assert report.overall.metrics["extraction_failure_rate"].sample_count == (
        2 * len(dataset.cases)
    )
    assert {group.dimension for group in report.groups} >= {
        "overall",
        "attribute_set",
        "product_profile",
        "attribute_header",
        "model",
        "analysis_mode",
    }


def test_variant_leakage_and_threshold_routing_are_deterministic() -> None:
    dataset = load_dataset(DATASET)

    def leaking(config, case):
        prediction = engineering_echo_prediction(config, case)
        if case.case_id == "topwear-blue-variant":
            return prediction.model_copy(
                update={"values": {**prediction.values, "attributes__color": "Red"}}
            )
        return prediction

    report = compare_models(dataset, (_model("a"), _model("b")), leaking)
    color = next(
        group for group in report.groups
        if group.dimension == "attribute_header" and group.key == "attributes__color"
    )
    assert color.metrics["variant_leakage_rate"].value == 1
    assert color.metrics["variant_leakage_rate"].sample_count == 2

    pending = ThresholdConfiguration(
        version="1",
        approval_status=ApprovalStatus.PENDING,
        default_policy=ThresholdPolicy.REVIEW_REQUIRED,
    )
    assert all(
        decision.policy == ThresholdPolicy.REVIEW_REQUIRED
        for decision in route_threshold_policies(report, pending)
    )

    approved = ThresholdConfiguration(
        version="1",
        approval_status=ApprovalStatus.APPROVED,
        default_policy=ThresholdPolicy.REVIEW_REQUIRED,
        fields={
            "attributes__color": FieldThreshold(
                maximum_variant_leakage_rate=0,
                failure_policy=ThresholdPolicy.EXPLICIT_INPUT_ONLY,
            )
        },
    )
    decision = next(
        item
        for item in route_threshold_policies(report, approved)
        if item.header == "attributes__color"
    )
    assert decision.policy == ThresholdPolicy.EXPLICIT_INPUT_ONLY
    assert "variant_leakage_rate=1.0000" in decision.reasons[0]


def test_distinct_models_are_required_for_comparison() -> None:
    with pytest.raises(ValueError, match="two distinct"):
        compare_models(load_dataset(DATASET), (_model("same"), _model("same")), lambda *_: None)


def test_resource_limits_and_costs_validate_boundaries(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    limits = ResourceLimits.from_env(
        {
            "FASHION_CMS_WORKBOOK_ROWS": "2",
            "FASHION_CMS_MODEL_CONCURRENCY": "2",
            "FASHION_CMS_MAXIMUM_ESTIMATED_COST": "3.50",
        }
    )
    assert (limits.workbook_rows, limits.model_concurrency) == (2, 2)
    assert any(row["Limit"] == "retention_days" for row in limits.health_rows())
    with pytest.raises(ValidationError):
        ResourceLimits(workbook_rows=0)
    with pytest.raises(ValidationError, match="at least workbook_bytes"):
        ResourceLimits(workbook_bytes=2, workbook_expanded_bytes=1)

    pricing = ModelPricing(
        model_id="configured-model",
        currency="USD",
        effective_date=date(2026, 1, 1),
        source="user-approved price sheet",
        input_per_million="2",
        output_per_million="4",
        image_pricing_method="PER_IMAGE",
        image_rate="0.01",
        maximum_input_tokens_per_request=100,
        maximum_output_tokens_per_request=50,
    )
    assert usage_cost(
        pricing, {"input_tokens": 1_000_000, "output_tokens": 500_000}, image_count=2
    ) == Decimal("4.02")
    assert maximum_job_cost(pricing, request_count=2, image_count=3) == Decimal("0.0308")
    assert usage_cost(None, {"input_tokens": 1, "output_tokens": 1}) is None
    assert usage_cost(pricing, {"input_tokens": 1}) is None

    workbook = Workbook()
    workbook.active.append(
        [
            "sku",
            "base_code",
            "attributes__lulu_ean",
            "attributes__shipping_weight",
            "model_code_input_data",
        ]
    )
    workbook.active.append(["AB", "B", "1", 0, "x"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    monkeypatch.setenv("FASHION_CMS_CELL_CHARACTERS", "1")
    result = parse_input_workbook(output.getvalue())
    assert not result.ready
    assert any("configured 1 character limit" in issue.message for issue in result.issues)


def test_catalog_usage_is_retained_for_cost_reporting() -> None:
    registry = load_registry(ROOT / "config" / "attribute_registry.xlsx")
    result = generate_catalog_content(
        {
            "attributes__brand": "Acme",
            "attributes__product_type": "T-Shirt",
            "attributes__material": "Cotton",
        },
        registry,
        fake_catalog_client(),
        model="phase6-fake",
    )
    assert result.request_count == 1
    assert result.retry_count == 0
    assert result.usage == {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}


def test_release_gate_report_never_promotes_missing_or_blocked_evidence() -> None:
    blocked = build_report(
        {
            "evaluation_dataset_approval": (
                GateStatus.BLOCKED_USER_DECISION,
                "Engineering fixture only.",
                "tests/fixtures/phase8_engineering_dataset.json",
                "Human approval is required.",
            )
        }
    )
    assert len(blocked.results) == len(GATES)
    assert blocked.verdict == "BLOCKED"
    assert not blocked.production_ready

    passed = build_report(
        {
            gate_id: (GateStatus.PASS, "verified", None, None)
            for gate_id, _, _ in GATES
        }
    )
    assert passed.production_ready
    assert passed.verdict == "READY_FOR_USER_ACCEPTANCE"


def test_committed_release_report_is_complete_and_blocked() -> None:
    report = load_report(ROOT / "docs" / "releases" / "0.1.0-rc1" / "release-gates.json")
    assert report.verdict == "BLOCKED"
    assert not report.production_ready
    assert any(result.status == GateStatus.BLOCKED_USER_DECISION for result in report.results)
    assert next(result for result in report.results if result.gate_id == "model_comparison").status == (
        GateStatus.NOT_RUN
    )


def test_workbook_external_content_is_rejected() -> None:
    workbook = Workbook()
    workbook.active.append(["sku"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    modified = BytesIO()
    with ZipFile(BytesIO(output.getvalue())) as source, ZipFile(modified, "w") as target:
        for member in source.infolist():
            target.writestr(member, source.read(member))
        target.writestr("xl/externalLinks/externalLink1.xml", b"<externalLink/>")

    issue = preflight_xlsx(modified.getvalue())
    assert issue is not None
    assert issue.code == "EXTERNAL_WORKBOOK_CONTENT"


@pytest.mark.parametrize(
    ("archive", "code"),
    [
        (_zip(("nested.zip", _zip(("SKU-1.png", b"x")))), "NESTED_ARCHIVE"),
        (
            _zip(
                ("one/SKU-1.png", _image_bytes()),
                ("two/SKU-1.png", _image_bytes()),
            ),
            "ARCHIVE_FILENAME_COLLISION",
        ),
    ],
)
def test_nested_archives_and_flat_filename_collisions_are_rejected(
    archive: bytes, code: str
) -> None:
    result = parse_uploaded_images((("images.zip", archive),), ("SKU",))
    assert not result.ready
    assert code in {issue.code for issue in result.issues}


def test_zip_symlinks_are_rejected() -> None:
    link = ZipInfo("SKU-1.png")
    link.create_system = 3
    link.external_attr = (stat.S_IFLNK | 0o777) << 16
    result = parse_uploaded_images(
        (("images.zip", _zip((link, b"target"))),),
        ("SKU",),
    )
    assert not result.ready
    assert "ARCHIVE_SYMLINK" in {issue.code for issue in result.issues}


@pytest.mark.filterwarnings("ignore:Saving I mode images")
def test_animated_unsupported_mode_and_excessive_dimensions_are_rejected() -> None:
    first = Image.new("RGB", (2, 2), "red")
    second = Image.new("RGB", (2, 2), "blue")
    animated = BytesIO()
    first.save(
        animated,
        format="WEBP",
        save_all=True,
        append_images=[second],
        duration=10,
        loop=0,
    )
    first.close()
    second.close()
    mode_i = _image_bytes("PNG", size=(2, 2), mode="I")
    result = parse_uploaded_images(
        (("SKU-1.webp", animated.getvalue()), ("SKU-2.png", mode_i)),
        ("SKU",),
    )
    assert {"MULTIFRAME_IMAGE", "UNSUPPORTED_IMAGE_MODE"} <= {
        issue.code for issue in result.issues
    }

    dimensions = parse_uploaded_images(
        (("SKU-1.png", _image_bytes()),),
        ("SKU",),
        limits=ResourceLimits(image_dimension=10),
    )
    assert "IMAGE_DIMENSIONS_TOO_LARGE" in {issue.code for issue in dimensions.issues}


def test_retry_after_and_total_deadline_are_bounded() -> None:
    calls = 0
    delays: list[float] = []

    def handler(_request: httpx.Request) -> httpx.Response:
        nonlocal calls
        calls += 1
        if calls == 1:
            return httpx.Response(429, headers={"Retry-After": "2"})
        return httpx.Response(
            200,
            headers={"Content-Type": "image/png"},
            content=_image_bytes(),
        )

    request = ImageUrlRequest(row_number=2, sku="SKU", ordinal=1, source_url=PUBLIC_URL)
    result = download_images(
        (request,),
        settings=DownloadSettings(retry_count=1),
        transport=httpx.MockTransport(handler),
        resolver=lambda *_: ("93.184.216.34",),
        sleeper=delays.append,
    )
    assert result.report[0].result == DownloadResult.SUCCESS
    assert delays == [2]

    ticks = iter((0.0, 1.0, 2.0, 3.0))
    deadline = download_images(
        (request,),
        settings=DownloadSettings(total_deadline_seconds=0.5),
        transport=httpx.MockTransport(lambda _request: pytest.fail("deadline must block fetch")),
        resolver=lambda *_: ("93.184.216.34",),
        clock=lambda: next(ticks),
    )
    assert deadline.report[0].error_message == "Total request deadline was exceeded."


def test_secret_redaction_covers_headers_tokens_credentials_and_named_values() -> None:
    secret = "private-value-123"
    message = sanitize_error(
        "Authorization: Bearer abcdefghijk api_key=another-secret "
        "https://user:password@example.com sk-abcdefghijkl " + secret,
        (secret,),
    )
    assert "another-secret" not in message
    assert "password" not in message
    assert "sk-abcdefghijkl" not in message
    assert secret not in message
    assert message.count("[redacted]") >= 4


def test_cleanup_is_root_scoped_active_safe_dry_run_and_idempotent(tmp_path: Path) -> None:
    root = tmp_path / "data"
    active = root / "active" / "artifact.bin"
    expired = root / "expired"
    active.parent.mkdir(parents=True)
    expired.mkdir()
    active.write_bytes(b"active")
    (expired / "artifact.bin").write_bytes(b"expired")

    dry = cleanup_paths(root, (expired, active.parent, tmp_path), active=(active,), dry_run=True)
    assert dry.deleted == ("expired",)
    assert dry.skipped_active == ("active",)
    assert str(tmp_path) in dry.refused
    assert expired.exists()

    deleted = cleanup_paths(root, (expired,), dry_run=False)
    assert deleted.deleted == ("expired",)
    assert not expired.exists()
    assert cleanup_paths(root, (expired,), dry_run=False).deleted == ()
    assert durable_cleanup_disabled().disabled_reason

    symlink = tmp_path / "root-link"
    symlink.symlink_to(root, target_is_directory=True)
    with pytest.raises(ValueError, match="non-symlink"):
        cleanup_paths(symlink, (), dry_run=True)


def test_bounded_concurrency_cancellation_and_resume_preserve_completed_work(
    tmp_path: Path,
) -> None:
    database = JobDatabase(tmp_path / "concurrency.sqlite3")
    service = JobService(database)
    job_id = service.create_job(
        _rows(6), attribute_set="topwear", registry_version="registry"
    )
    active = 0
    maximum = 0
    lock = threading.Lock()

    def bounded(item):
        nonlocal active, maximum
        with lock:
            active += 1
            maximum = max(maximum, active)
        time.sleep(0.01)
        with lock:
            active -= 1
        return fake_extract(item)

    completed = service.run_job(job_id, bounded, limits=ResourceLimits(model_concurrency=2))
    assert completed.status == JobStatus.COMPLETED
    assert 1 < maximum <= 2

    cancelled_id = service.create_job(
        _rows(3), attribute_set="topwear", registry_version="registry", model_identifier="other"
    )
    first_calls: list[str] = []

    def cancel_after_first(item):
        first_calls.append(item.key)
        service.request_cancellation(cancelled_id)
        return fake_extract(item)

    cancelled = service.run_job(
        cancelled_id,
        cancel_after_first,
        limits=ResourceLimits(model_concurrency=1),
    )
    items = database.list_work_items(cancelled_id)
    assert cancelled.status == JobStatus.PARTIAL_FAILURE
    assert cancelled.cancel_requested
    assert [item.status for item in items].count(WorkItemStatus.COMPLETED) == 1
    assert [item.status for item in items].count(WorkItemStatus.PENDING) == 2

    resumed_calls: list[str] = []

    def resumed(item):
        resumed_calls.append(item.key)
        return fake_extract(item)

    resumed_job = service.resume_job(cancelled_id, resumed)
    assert resumed_job.status == JobStatus.COMPLETED
    assert not resumed_job.cancel_requested
    assert len(first_calls) == 1
    assert len(resumed_calls) == 2


def test_persisted_call_circuit_and_database_backup(tmp_path: Path) -> None:
    source = tmp_path / "jobs.sqlite3"
    database = JobDatabase(source)
    service = JobService(database)
    job_id = service.create_job(
        _rows(1), attribute_set="topwear", registry_version="registry"
    )
    item = database.list_work_items(job_id)[0]
    database.transition_job(job_id, JobStatus.RUNNING)
    database.mark_item_running(job_id, item.key)
    assert database.claim_model_call(job_id, 1, item_key=item.key)
    assert not database.claim_model_call(job_id, 1, item_key=item.key, retry=True)
    stored = database.list_work_items(job_id)[0]
    assert (database.get_job(job_id).attempted_model_calls, stored.attempted_model_calls) == (1, 1)
    assert stored.provider_retry_count == 0

    backup = database.backup(tmp_path / "backup" / "jobs.sqlite3")
    restored = JobDatabase(backup)
    assert restored.get_job(job_id).attempted_model_calls == 1
    with pytest.raises(FileExistsError):
        database.backup(backup)


def test_partial_qc_lists_incomplete_rows_as_literal_text() -> None:
    workbook = build_qc_report(
        (),
        incomplete_rows=(("=BAD", "FAILED", "+unsafe"),),
    )
    loaded = load_workbook(BytesIO(workbook), data_only=False)
    try:
        sheet = loaded["Incomplete"]
        assert sheet["A2"].data_type == "s"
        assert sheet["A2"].value == "'=BAD"
        assert sheet["C2"].value == "'+unsafe"
    finally:
        loaded.close()


def test_evaluation_artifact_matches_the_frozen_engineering_fixture() -> None:
    artifact = json.loads(
        (ROOT / "docs" / "releases" / "0.1.0-rc1" / "evaluation-report.json").read_text()
    )
    dataset = load_dataset(DATASET)
    assert artifact["dataset_version"] == dataset.dataset_version
    assert artifact["dataset_fingerprint"] == dataset.fingerprint
    assert artifact["sample_counts"]["cases"] == len(dataset.cases)
    assert artifact["evidence_class"] == "ENGINEERING_FIXTURE_ONLY"
    assert artifact["live_model_comparison"] == "NOT_RUN"
