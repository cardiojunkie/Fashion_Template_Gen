from __future__ import annotations

import itertools
import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fashion_cms.catalog_service import (
    CatalogBullet,
    CatalogContent,
    CatalogWireOutput,
    ContentValidationError,
    ExportBlockedError,
    KeywordGroup,
    SkuCatalog,
    TOPWEAR_HEADERS,
    build_qc_report,
    build_topwear_title,
    build_topwear_workbook,
    fake_catalog_client,
    generate_catalog_content,
    sanitize_excel_text,
    validate_catalog_output,
)
from fashion_cms.llm_service import LLMResponse
from fashion_cms.models import InputRow
from fashion_cms.normalization import MatchMethod
from fashion_cms.registry import load_registry
from fashion_cms.review import (
    ProposalStatus,
    ReviewAction,
    ReviewItem,
    SourcePriority,
)
from fashion_cms.topwear_extraction import Confidence


ROOT = Path(__file__).parents[1]


@pytest.fixture(scope="module")
def registry():
    return load_registry(ROOT / "config" / "attribute_registry.xlsx")


def reviewed_item(
    sku: str,
    header: str,
    value: str,
    registry,
    *,
    image_color: bool = False,
    resolved: bool = True,
) -> ReviewItem:
    return ReviewItem(
        job_id="job",
        sku=sku,
        base_code="BASE",
        product_profile="topwear_mvp",
        header=header,
        supplied_value=None if image_color else value,
        raw_value=value,
        normalized_value=value.casefold(),
        proposed_value=value,
        matching_method=(
            MatchMethod.EXACT_CANONICAL
            if header in {"attributes__color", "attributes__fit_type"}
            else MatchMethod.FREE_TEXT
        ),
        alias_used=None,
        fuzzy_suggestion=None,
        fuzzy_score=None,
        evidence_type="image" if image_color else "structured_input",
        evidence_references=("0001-1",) if image_color else (f"input:{sku}:{header}",),
        confidence=Confidence.HIGH,
        source_priority=(SourcePriority.IMAGE if image_color else SourcePriority.STRUCTURED_INPUT),
        conflict=None,
        warning=(
            f"Color inferred from image using broad value: {value}" if image_color else None
        ),
        proposal_status=ProposalStatus.PROPOSED,
        image_inferred_color=image_color,
        requires_review=True,
        review_action=ReviewAction.ACCEPT if resolved else None,
        final_value=value if resolved else None,
        decision_valid=True,
        registry_version=registry.fingerprint,
        prompt_version="prompt",
        schema_version="schema",
        model="fake",
    )


def content(*bullets: str) -> CatalogContent:
    padded = tuple((list(bullets) + [""] * 6)[:6])
    return CatalogContent(
        keywords="Acme, T-Shirt",
        bullets=padded,
        keyword_source_headers=("attributes__brand", "attributes__product_type"),
        bullet_source_headers=tuple(("attributes__product_type",) for _ in bullets),
        model="fake",
    )


def test_complete_title_order_and_duplicate_model_handling() -> None:
    assert build_topwear_title(
        brand="Acme",
        series_name="Core",
        material="Cotton",
        product_type="T-Shirt",
        size="M",
        color="Blue",
        model_number="M1",
    ) == "Acme Core Cotton T-Shirt, M, Blue, M1"
    assert build_topwear_title(
        brand="Acme",
        series_name="Core M1",
        product_type="T-Shirt",
        model_number="M1",
    ) == "Acme Core M1 T-Shirt"
    assert build_topwear_title(brand="Blue", color="Blue") == "Blue"


def test_every_missing_title_component_combination_is_well_formed() -> None:
    values = ("Brand", "Series", "Cotton", "T-Shirt", "M", "Blue", "M1")
    for included in itertools.product((False, True), repeat=len(values)):
        title = build_topwear_title(
            **dict(
                zip(
                    (
                        "brand",
                        "series_name",
                        "material",
                        "product_type",
                        "size",
                        "color",
                        "model_number",
                    ),
                    (value if keep else None for value, keep in zip(values, included)),
                    strict=True,
                )
            )
        )
        assert "  " not in title
        assert ",," not in title
        assert not title.startswith(",")
        assert not title.endswith(",")
        assert "Unknown" not in title


def test_fake_catalog_request_uses_only_accepted_facts_and_no_images(registry) -> None:
    accepted = {
        "attributes__brand": "Acme",
        "attributes__product_type": "T-Shirt",
        "attributes__material": "Cotton",
        "attributes__color": "Blue",
    }
    generated = generate_catalog_content(
        accepted, registry, fake_catalog_client(), model="phase6-fake"
    )

    assert generated.keywords == "Acme, T-Shirt, Cotton, Blue"
    assert generated.bullets[:3] == (
        "Made with Cotton.",
        "T-Shirt construction.",
        "",
    )
    assert generated.warnings
    assert all("input_image" not in bullet for bullet in generated.bullets)


@pytest.mark.parametrize(
    ("text", "sources", "message"),
    (
        ("Premium Cotton fabric.", ("attributes__material",), "promotional"),
        ("Blue color.", ("attributes__color",), "Color-, size-, or model-only"),
        ("Model year 2026.", ("attributes__material",), "Model year"),
        ("Information is missing.", ("attributes__material",), "missing-information"),
        ("input:0001 says Cotton.", ("attributes__material",), "Internal evidence"),
    ),
)
def test_unsupported_or_internal_bullets_are_rejected(
    registry, text: str, sources: tuple[str, ...], message: str
) -> None:
    wire = CatalogWireOutput(
        keyword_groups=(),
        bullets=(CatalogBullet(text=text, source_headers=sources),),
    )
    with pytest.raises(ContentValidationError, match=message):
        validate_catalog_output(
            wire,
            {"attributes__material": "Cotton", "attributes__color": "Blue"},
            registry,
        )


def test_repeated_bullets_openings_and_unaccepted_sources_are_rejected(registry) -> None:
    wire = CatalogWireOutput(
        keyword_groups=(
            KeywordGroup(text="Cotton", source_headers=("attributes__material",)),
            KeywordGroup(text="cotton", source_headers=("attributes__material",)),
        ),
        bullets=(
            CatalogBullet(text="Cotton fabric.", source_headers=("attributes__material",)),
            CatalogBullet(text="Cotton fabric.", source_headers=("attributes__material",)),
        ),
    )
    with pytest.raises(ContentValidationError, match="Repeated"):
        validate_catalog_output(
            wire, {"attributes__material": "Cotton"}, registry
        )


def test_generation_retries_validation_once(registry) -> None:
    valid = json.dumps(
        {
            "keyword_groups": [
                {"text": "Cotton", "source_headers": ["attributes__material"]}
            ],
            "bullets": [
                {"text": "Cotton fabric.", "source_headers": ["attributes__material"]}
            ],
        }
    )

    class Client:
        calls = 0

        def create(self, request):
            self.calls += 1
            return LLMResponse(
                model="fake",
                status="completed",
                output_text="not json" if self.calls == 1 else valid,
            )

    client = Client()
    generated = generate_catalog_content(
        {"attributes__material": "Cotton"}, registry, client, model="fake"
    )
    assert client.calls == 2
    assert generated.bullets[0] == "Cotton fabric."


def test_excel_sanitization_preserves_punctuation_and_neutralizes_formulas() -> None:
    assert sanitize_excel_text("Crew-neck T-Shirt") == "Crew-neck T-Shirt"
    assert sanitize_excel_text("=HYPERLINK('bad')") == "'=HYPERLINK('bad')"


def test_export_has_exact_headers_identifiers_and_color_highlighting(registry) -> None:
    rows = (
        InputRow(
            row_number=2,
            sku="0001",
            base_code="0010",
            attributes__lulu_ean="0000009",
            attributes__shipping_weight=0,
        ),
        InputRow(
            row_number=3,
            sku="0002",
            base_code="0010",
            attributes__lulu_ean="0000010",
            attributes__shipping_weight="1.5",
        ),
    )
    items = (
        reviewed_item("0001", "attributes__brand", "=Acme", registry),
        reviewed_item(
            "0001", "attributes__color", "Blue", registry, image_color=True
        ),
        reviewed_item("0002", "attributes__brand", "Acme", registry),
        reviewed_item("0002", "attributes__color", "Red", registry),
    )
    catalogs = {
        "0001": SkuCatalog(sku="0001", title="Acme Shirt, Blue", content=content()),
        "0002": SkuCatalog(sku="0002", title="Acme Shirt, Red", content=content()),
    }
    exported = build_topwear_workbook(rows, items, catalogs, registry)
    workbook = load_workbook(BytesIO(exported), data_only=False)
    worksheet = workbook["CMS Upload"]
    headers = tuple(cell.value for cell in worksheet[1])
    color_column = headers.index("attributes__color") + 1
    brand_column = headers.index("attributes__brand") + 1

    assert headers == TOPWEAR_HEADERS
    assert worksheet.max_row == 3
    assert worksheet.cell(2, 1).value == "0001"
    assert worksheet.cell(2, 3).value == "0000009"
    assert worksheet.cell(2, 4).value == 0
    assert worksheet.cell(2, brand_column).value == "'=Acme"
    assert str(worksheet.cell(2, color_column).fill.fgColor.rgb).endswith("FFFF00")
    assert worksheet.cell(3, color_column).fill.fill_type is None
    assert workbook.sheetnames == ["CMS Upload"]
    workbook.close()

    qc = load_workbook(BytesIO(build_qc_report(items)), data_only=False)
    qc_sheet = qc["Topwear QC"]
    notes = [cell.value for cell in qc_sheet[1]]
    note_column = notes.index("Image-inference note") + 1
    assert "Color inferred from image" in qc_sheet.cell(2 + 1, note_column).value
    qc.close()


def test_invalid_enum_or_unresolved_review_blocks_export(registry) -> None:
    row = InputRow(row_number=2, sku="0001")
    catalog = {"0001": SkuCatalog(sku="0001", title="", content=content())}
    invalid = reviewed_item("0001", "attributes__color", "Navy Blue", registry)
    unresolved = reviewed_item(
        "0001", "attributes__brand", "Acme", registry, resolved=False
    )

    with pytest.raises(ExportBlockedError, match="permitted"):
        build_topwear_workbook((row,), (invalid,), catalog, registry)
    with pytest.raises(ExportBlockedError, match="unresolved"):
        build_topwear_workbook((row,), (unresolved,), catalog, registry)
