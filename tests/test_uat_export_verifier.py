from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from fashion_cms.excel_service import build_blank_cms_workbook, parse_input_workbook
from fashion_cms.registry import load_registry
from uat.scripts.verify_exports import main, verify_export


ROOT = Path(__file__).parents[1]
REGISTRY = ROOT / "config" / "attribute_registry.xlsx"


def _input_workbook(path: Path) -> tuple:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        (
            "sku",
            "base_code",
            "attributes__lulu_ean",
            "attributes__shipping_weight",
            "model_code_input_data",
        )
    )
    sheet.append(("000123", "000BASE", "0000000123456", 0.25, "Structural test"))
    for cell in sheet[2][:3]:
        cell.number_format = "@"
    workbook.save(path)
    workbook.close()
    parsed = parse_input_workbook(path.read_bytes(), path.name)
    assert parsed.ready
    return parsed.rows


def _export(path: Path, input_path: Path, attribute_set: str = "topwear") -> None:
    registry = load_registry(REGISTRY)
    path.write_bytes(
        build_blank_cms_workbook(
            _input_workbook(input_path), registry.mappings_by_set[attribute_set]
        )
    )


def _edit(path: Path, callback) -> None:
    workbook = load_workbook(path)
    callback(workbook.active)
    workbook.save(path)
    workbook.close()


def test_verifier_passes_exact_export_and_writes_machine_report(
    tmp_path: Path, capsys
) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    report_path = tmp_path / "report.json"
    _export(export_path, input_path)

    assert (
        main(
            [
                "--attribute-set",
                "topwear",
                "--input-workbook",
                str(input_path),
                "--report-json",
                str(report_path),
                str(export_path),
            ]
        )
        == 0
    )
    assert "PASS:" in capsys.readouterr().out
    assert json.loads(report_path.read_text())["status"] == "PASS"


def test_verifier_rejects_wrong_headers_internal_columns_and_row_count(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    _export(export_path, input_path)

    def corrupt(sheet) -> None:
        sheet.cell(1, 1, "internal_confidence")
        sheet.append(["UNEXPECTED"])

    _edit(export_path, corrupt)
    result = verify_export(
        export_path, attribute_set="topwear", input_workbook=input_path
    )

    assert result["status"] == "FAIL"
    assert {finding["code"] for finding in result["findings"]} >= {
        "HEADERS",
        "INTERNAL_COLUMNS",
        "MISSING_SKU",
    }


def test_verifier_rejects_non_text_or_changed_identifiers(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    _export(export_path, input_path)

    _edit(export_path, lambda sheet: sheet.cell(2, 3, 123456))
    result = verify_export(
        export_path, attribute_set="topwear", input_workbook=input_path
    )

    codes = {finding["code"] for finding in result["findings"]}
    assert {"IDENTIFIER_TEXT", "IDENTIFIER_VALUE"} <= codes


def test_verifier_rejects_invalid_enum_and_formula_like_output(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    _export(export_path, input_path)
    registry = load_registry(REGISTRY)
    headers = registry.mappings_by_set["topwear"]

    def corrupt(sheet) -> None:
        sheet.cell(2, headers.index("attributes__color") + 1, "Purple")
        sheet.cell(2, headers.index("attributes__other_information") + 1, "=1+1")

    _edit(export_path, corrupt)
    result = verify_export(
        export_path, attribute_set="topwear", input_workbook=input_path
    )

    codes = {finding["code"] for finding in result["findings"]}
    assert {"INVALID_ENUM", "UNSAFE_FORMULA"} <= codes


def test_verifier_rejects_profile_inapplicable_accessory_value(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    _export(export_path, input_path, "mens_accessories")
    registry = load_registry(REGISTRY)
    headers = registry.mappings_by_set["mens_accessories"]
    _edit(
        export_path,
        lambda sheet: sheet.cell(2, headers.index("attributes__bag_type") + 1, "Backpack"),
    )

    result = verify_export(
        export_path,
        attribute_set="mens_accessories",
        input_workbook=input_path,
        profile="watches",
    )

    assert any(
        finding["code"] == "PROFILE_INAPPLICABLE" for finding in result["findings"]
    )


def test_verifier_does_not_modify_user_workbook(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    _export(export_path, input_path)
    before = export_path.read_bytes()

    verify_export(export_path, attribute_set="topwear", input_workbook=input_path)

    assert export_path.read_bytes() == before


def test_verifier_reuses_secure_workbook_preflight(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    export_path = tmp_path / "export.xlsx"
    _export(export_path, input_path)
    unsafe_path = tmp_path / "unsafe.xlsx"
    with ZipFile(export_path) as source, ZipFile(unsafe_path, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            target.writestr(member, source.read(member))
        target.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")

    result = verify_export(
        unsafe_path, attribute_set="topwear", input_workbook=input_path
    )

    assert result["status"] == "FAIL"
    assert result["findings"] == [
        {
            "code": "EXTERNAL_WORKBOOK_CONTENT",
            "message": "Workbook external links and data connections are not supported.",
        }
    ]
