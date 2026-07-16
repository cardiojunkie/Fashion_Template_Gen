from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from fashion_cms.excel_service import REQUIRED_COLUMNS, parse_input_workbook
from fashion_cms.registry import load_registry


ROOT = Path(__file__).parents[1]
UAT = ROOT / "uat"


def test_expected_headers_are_generated_from_the_active_registry() -> None:
    registry = load_registry(ROOT / "config" / "attribute_registry.xlsx")
    expected = json.loads((UAT / "expected_headers.json").read_text())

    assert expected["registry_fingerprint"] == registry.fingerprint
    assert {
        row["canonical_id"]: (row["display_name"], row["header_count"], row["ordered_headers"])
        for row in expected["attribute_sets"]
    } == {
        set_id: (
            next(
                row.attribute_set_name
                for row in registry.attribute_sets
                if row.attribute_set_id == set_id
            ),
            len(headers),
            list(headers),
        )
        for set_id, headers in registry.mappings_by_set.items()
    }


def test_manual_checklist_has_required_sheets_columns_and_status_dropdowns() -> None:
    workbook = load_workbook(UAT / "manual_uat_checklist.xlsx", read_only=False)
    try:
        assert workbook.sheetnames == [
            "Instructions",
            "Environment",
            "Phase Audit",
            "CMS Generator",
            "Variant Testing",
            "Review and Catalog Copy",
            "Attribute Set Exports",
            "Men’s Accessories",
            "Image Downloader",
            "Job Recovery",
            "Security Checks",
            "NVIDIA Connection",
            "Defects",
            "User Sign-Off",
        ]
        expected_columns = (
            "Test ID",
            "Module",
            "Preconditions",
            "Input/Test Data",
            "Steps",
            "Expected Result",
            "Actual Result",
            "Status",
            "Evidence/Screenshot",
            "Defect ID",
            "Notes",
        )
        for sheet in workbook.worksheets[2:]:
            assert tuple(cell.value for cell in sheet[1]) == expected_columns
            assert len(sheet.data_validations.dataValidation) == 1
            assert "NOT STARTED,PASS,FAIL,BLOCKED,NOT APPLICABLE" in str(
                sheet.data_validations.dataValidation[0].formula1
            )
    finally:
        workbook.close()


def test_structural_inputs_are_text_safe_and_negative_fixtures_are_actionable() -> None:
    registry = load_registry(ROOT / "config" / "attribute_registry.xlsx")
    for set_id in registry.mappings_by_set:
        path = UAT / "inputs" / f"{set_id}_structural.xlsx"
        result = parse_input_workbook(path.read_bytes(), path.name)
        assert result.ready
        assert result.rows[0].sku == "000123"
        assert result.rows[0].attributes__lulu_ean == "0000000123456"

    duplicate = UAT / "inputs" / "duplicate_sku.xlsx"
    duplicate_result = parse_input_workbook(duplicate.read_bytes(), duplicate.name)
    assert not duplicate_result.ready
    assert any(issue.code == "DUPLICATE_SKU" for issue in duplicate_result.issues)

    missing = UAT / "inputs" / "missing_required_column.xlsx"
    missing_result = parse_input_workbook(missing.read_bytes(), missing.name)
    assert not missing_result.ready
    assert any(issue.code == "MISSING_COLUMNS" for issue in missing_result.issues)

    downloader = load_workbook(
        UAT / "inputs" / "image_downloader_uat.xlsx", read_only=True, data_only=False
    )
    try:
        sheet = downloader.active
        assert tuple(cell.value for cell in sheet[1]) == (
            "sku",
            "Image 1",
            "Image 2",
            "Image 3",
            "Image 4",
        )
        assert all(cell.value is None for row in sheet.iter_rows(min_row=2) for cell in row[1:])
    finally:
        downloader.close()

    assert tuple(REQUIRED_COLUMNS) == (
        "sku",
        "base_code",
        "attributes__lulu_ean",
        "attributes__shipping_weight",
        "input_data",
    )
