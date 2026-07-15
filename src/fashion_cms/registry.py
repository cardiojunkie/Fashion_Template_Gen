from __future__ import annotations

import argparse
import hashlib
import unicodedata
from collections import defaultdict
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, ValidationError


class DataType(StrEnum):
    ENUM = "ENUM"
    FREE_TEXT = "FREE_TEXT"
    INTEGER = "INTEGER"
    DECIMAL = "DECIMAL"
    BOOLEAN = "BOOLEAN"
    SYSTEM_COPY = "SYSTEM_COPY"
    GENERATED_TEXT = "GENERATED_TEXT"


class Scope(StrEnum):
    SYSTEM = "SYSTEM"
    SKU = "SKU"
    VARIANT = "VARIANT"
    STYLE = "STYLE"
    JOB = "JOB"


class EvidencePolicy(StrEnum):
    SYSTEM_COPY = "SYSTEM_COPY"
    EXPLICIT_TEXT_ONLY = "EXPLICIT_TEXT_ONLY"
    VISUAL_OR_TEXT = "VISUAL_OR_TEXT"
    DERIVED_BUSINESS_RULE = "DERIVED_BUSINESS_RULE"
    GENERATED_CONTENT = "GENERATED_CONTENT"


class RegistryRow(BaseModel):
    model_config = ConfigDict(frozen=True, str_strip_whitespace=True, extra="ignore")


class AttributeSetRow(RegistryRow):
    attribute_set_id: str = Field(min_length=1)
    attribute_set_name: str = Field(min_length=1)
    position: int = Field(gt=0)
    header: str = Field(min_length=1)
    required: bool


class AttributeDefinition(RegistryRow):
    header: str = Field(min_length=1)
    data_type: DataType
    scope: Scope
    evidence_policy: EvidencePolicy
    nullable: bool
    description: str = Field(min_length=1)
    unit_or_format: str | None = None


class PermittedValuesRow(RegistryRow):
    attribute_header: str = Field(min_length=1)
    data_type: DataType
    values: tuple[str, ...]


class ValueAlias(RegistryRow):
    attribute_header: str = Field(min_length=1)
    alias: str = Field(min_length=1)
    canonical_value: str = Field(min_length=1)
    active: bool


class ProductProfile(RegistryRow):
    attribute_set_id: str = Field(min_length=1)
    product_type: str | None = Field(default=None, min_length=1)
    profile_id: str = Field(min_length=1)
    header: str = Field(min_length=1)
    applicable: bool


class Registry(BaseModel):
    model_config = ConfigDict(frozen=True)

    attribute_sets: tuple[AttributeSetRow, ...]
    definitions: tuple[AttributeDefinition, ...]
    permitted_values: tuple[PermittedValuesRow, ...]
    aliases: tuple[ValueAlias, ...]
    profiles: tuple[ProductProfile, ...]
    mappings_by_set: dict[str, tuple[str, ...]]
    definitions_by_header: dict[str, AttributeDefinition]
    permitted_values_by_header: dict[str, tuple[str, ...]]
    aliases_by_header: dict[str, dict[str, str]]
    profiles_by_id: dict[tuple[str, str], tuple[ProductProfile, ...]]
    configuration_issues_by_set: dict[str, tuple[str, ...]]
    fingerprint: str


class RegistryValidationError(ValueError):
    def __init__(self, errors: list[str]) -> None:
        self.errors = tuple(errors)
        super().__init__("Invalid attribute registry:\n- " + "\n- ".join(errors))


REQUIRED_COLUMNS = {
    "Attribute_Sets": (
        "attribute_set_id",
        "attribute_set_name",
        "position",
        "header",
        "required",
    ),
    "Attribute_Definitions": (
        "header",
        "data_type",
        "scope",
        "evidence_policy",
        "nullable",
        "description",
        "unit_or_format",
    ),
    "Permitted_Values": ("attribute_header", "data_type", "value_1"),
    "Value_Aliases": ("attribute_header", "alias", "canonical_value", "active"),
    "Product_Profiles": (
        "attribute_set_id",
        "product_type",
        "profile_id",
        "header",
        "applicable",
    ),
}

SYSTEM_HEADERS = {
    "sku",
    "base_code",
    "attributes__lulu_ean",
    "attributes__shipping_weight",
}
GENERATED_HEADERS = {
    "attributes__keywords",
    "name",
    "attributes__product_title",
    *(f"attributes__bullet_point_{number}" for number in range(1, 7)),
}
MANDATORY_ACCESSORY_PROFILES = (
    "bags_luggage",
    "caps_headwear",
    "watches",
    "eyewear",
    "belts_wallets_ties_other",
)
PROFILE_EXCLUDED_DATA_TYPES = {DataType.SYSTEM_COPY, DataType.GENERATED_TEXT}


def normalize_value(value: str) -> str:
    """Return the comparison form used for canonical values and aliases."""
    normalized = unicodedata.normalize("NFKC", value).casefold()
    return " ".join("".join(char if char.isalnum() else " " for char in normalized).split())


def _table_rows(worksheet: Any, sheet_name: str, errors: list[str]) -> list[tuple[int, dict]]:
    values = worksheet.iter_rows(values_only=True)
    first_row = next(values, None)
    if first_row is None:
        errors.append(f"{sheet_name}: missing header row")
        return []

    headers = [str(value).strip() if value is not None else "" for value in first_row]
    if len(headers) != len(set(headers)):
        errors.append(f"{sheet_name}: duplicate column header")
    missing = [column for column in REQUIRED_COLUMNS[sheet_name] if column not in headers]
    if missing:
        errors.append(f"{sheet_name}: missing columns {', '.join(missing)}")
        return []

    rows = []
    for row_number, row in enumerate(values, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        rows.append((row_number, dict(zip(headers, row, strict=False))))
    return rows


def _parse_rows(
    model: type[RegistryRow], rows: list[tuple[int, dict]], sheet_name: str, errors: list[str]
) -> list:
    parsed = []
    for row_number, row in rows:
        try:
            parsed.append(model.model_validate(row))
        except ValidationError as exc:
            details = "; ".join(error["msg"] for error in exc.errors(include_url=False))
            errors.append(f"{sheet_name} row {row_number}: {details}")
    return parsed


def _parse_permitted_values(
    rows: list[tuple[int, dict]], value_columns: list[str], errors: list[str]
) -> list[PermittedValuesRow]:
    expected_columns = [f"value_{number}" for number in range(1, len(value_columns) + 1)]
    if value_columns != expected_columns:
        errors.append("Permitted_Values: value columns must be consecutive from value_1")

    parsed = []
    for row_number, row in rows:
        raw_values = [row.get(column) for column in value_columns]
        first_blank = next(
            (
                index
                for index, value in enumerate(raw_values)
                if value is None or not str(value).strip()
            ),
            len(raw_values),
        )
        if any(value is not None and str(value).strip() for value in raw_values[first_blank:]):
            errors.append(f"Permitted_Values row {row_number}: canonical values contain a gap")
        values = tuple(str(value).strip() for value in raw_values[:first_blank])
        try:
            parsed.append(
                PermittedValuesRow.model_validate(
                    {
                        "attribute_header": row.get("attribute_header"),
                        "data_type": row.get("data_type"),
                        "values": values,
                    }
                )
            )
        except ValidationError as exc:
            details = "; ".join(error["msg"] for error in exc.errors(include_url=False))
            errors.append(f"Permitted_Values row {row_number}: {details}")
    return parsed


def _duplicates(values: list[str]) -> set[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        key = normalize_value(value)
        if key in seen:
            duplicates.add(value)
        seen.add(key)
    return duplicates


def _validate_registry(
    attribute_sets: list[AttributeSetRow],
    definitions: list[AttributeDefinition],
    permitted_values: list[PermittedValuesRow],
    aliases: list[ValueAlias],
    profiles: list[ProductProfile],
) -> list[str]:
    errors: list[str] = []
    mappings: dict[str, list[AttributeSetRow]] = defaultdict(list)
    for row in attribute_sets:
        mappings[row.attribute_set_id].append(row)

    for set_id, rows in mappings.items():
        names = {row.attribute_set_name for row in rows}
        if len(names) != 1:
            errors.append(f"Attribute_Sets: {set_id} has inconsistent names")
        duplicate_headers = _duplicates([row.header for row in rows])
        if duplicate_headers:
            errors.append(
                f"Attribute_Sets: {set_id} has duplicate headers {sorted(duplicate_headers)}"
            )
        positions = [row.position for row in rows]
        if sorted(positions) != list(range(1, len(rows) + 1)):
            errors.append(
                f"Attribute_Sets: {set_id} positions must be unique and consecutive from 1"
            )

    duplicate_definitions = _duplicates([row.header for row in definitions])
    if duplicate_definitions:
        errors.append(f"Attribute_Definitions: duplicate headers {sorted(duplicate_definitions)}")
    definitions_by_header = {row.header: row for row in definitions}
    mapped_headers = {row.header for row in attribute_sets}
    missing_definitions = sorted(mapped_headers - definitions_by_header.keys())
    if missing_definitions:
        errors.append(f"Attribute_Definitions: missing definitions {missing_definitions}")

    duplicate_value_rows = _duplicates([row.attribute_header for row in permitted_values])
    if duplicate_value_rows:
        errors.append(f"Permitted_Values: duplicate attribute rows {sorted(duplicate_value_rows)}")
    values_by_header = {row.attribute_header: row for row in permitted_values}
    missing_value_rows = sorted(mapped_headers - values_by_header.keys())
    if missing_value_rows:
        errors.append(f"Permitted_Values: missing attribute rows {missing_value_rows}")

    for row in permitted_values:
        definition = definitions_by_header.get(row.attribute_header)
        if definition is None:
            errors.append(f"Permitted_Values: unknown header {row.attribute_header}")
            continue
        if row.data_type != definition.data_type:
            errors.append(f"Permitted_Values: data type mismatch for {row.attribute_header}")
        duplicates = _duplicates(list(row.values))
        if duplicates:
            errors.append(
                f"Permitted_Values: {row.attribute_header} has duplicate canonical values "
                f"{sorted(duplicates)}"
            )
        if definition.data_type == DataType.ENUM and not row.values:
            errors.append(f"Permitted_Values: enum {row.attribute_header} has no canonical values")

    for header in SYSTEM_HEADERS & mapped_headers:
        definition = definitions_by_header.get(header)
        if definition and (
            definition.data_type != DataType.SYSTEM_COPY
            or definition.evidence_policy != EvidencePolicy.SYSTEM_COPY
        ):
            errors.append(f"Attribute_Definitions: system field {header} must use SYSTEM_COPY")
    for header in GENERATED_HEADERS & mapped_headers:
        definition = definitions_by_header.get(header)
        if definition and (
            definition.data_type != DataType.GENERATED_TEXT
            or definition.evidence_policy != EvidencePolicy.GENERATED_CONTENT
        ):
            errors.append(f"Attribute_Definitions: generated field {header} is misconfigured")

    alias_keys: dict[str, list[str]] = defaultdict(list)
    for alias in aliases:
        alias_keys[alias.attribute_header].append(alias.alias)
        if alias.attribute_header not in definitions_by_header:
            errors.append(f"Value_Aliases: unknown header {alias.attribute_header}")
            continue
        canonical_values = values_by_header.get(alias.attribute_header)
        canonical_keys = (
            {normalize_value(value) for value in canonical_values.values}
            if canonical_values
            else set()
        )
        if normalize_value(alias.canonical_value) not in canonical_keys:
            errors.append(
                f"Value_Aliases: alias {alias.alias!r} points to missing canonical value "
                f"{alias.canonical_value!r}"
            )
    for header, header_aliases in alias_keys.items():
        duplicates = _duplicates(header_aliases)
        if duplicates:
            errors.append(f"Value_Aliases: {header} has duplicate aliases {sorted(duplicates)}")

    profile_rows: dict[tuple[str, str], list[ProductProfile]] = defaultdict(list)
    seen_profile_headers: set[tuple[str, str, str]] = set()
    product_type_profiles: dict[tuple[str, str], set[str]] = defaultdict(set)
    for profile in profiles:
        rows = mappings.get(profile.attribute_set_id)
        if rows is None:
            errors.append(f"Product_Profiles: unknown attribute set {profile.attribute_set_id}")
            continue
        mapped_headers = {row.header for row in rows}
        if profile.header not in mapped_headers:
            errors.append(
                f"Product_Profiles: header {profile.header} is not in {profile.attribute_set_id}"
            )
            continue
        definition = definitions_by_header.get(profile.header)
        if definition and definition.data_type in PROFILE_EXCLUDED_DATA_TYPES:
            errors.append(f"Product_Profiles: {profile.header} cannot be profile-configured")

        normalized_profile_id = normalize_value(profile.profile_id)
        profile_key = (profile.attribute_set_id, normalized_profile_id)
        profile_rows[profile_key].append(profile)
        header_key = (*profile_key, profile.header)
        if header_key in seen_profile_headers:
            errors.append(
                "Product_Profiles: duplicate mapping for "
                f"{profile.attribute_set_id}/{profile.profile_id}/{profile.header}"
            )
        seen_profile_headers.add(header_key)
        if profile.product_type:
            product_type_profiles[
                (profile.attribute_set_id, normalize_value(profile.product_type))
            ].add(normalized_profile_id)

    for (set_id, normalized_profile_id), rows in profile_rows.items():
        profile_ids = {row.profile_id for row in rows}
        if len(profile_ids) != 1:
            errors.append(
                f"Product_Profiles: normalized duplicate profile IDs in {set_id}: "
                f"{sorted(profile_ids)}"
            )
        product_types = {
            normalize_value(row.product_type) if row.product_type else None for row in rows
        }
        if len(product_types) != 1:
            errors.append(
                f"Product_Profiles: {set_id}/{rows[0].profile_id} has inconsistent product types"
            )
        eligible_headers = {
            row.header
            for row in mappings[set_id]
            if (definition := definitions_by_header.get(row.header))
            and definition.data_type not in PROFILE_EXCLUDED_DATA_TYPES
        }
        configured_headers = {row.header for row in rows}
        missing_headers = sorted(eligible_headers - configured_headers)
        if missing_headers:
            errors.append(
                f"Product_Profiles: {set_id}/{rows[0].profile_id} is missing headers "
                f"{missing_headers}"
            )

    for (set_id, product_type), profile_ids in product_type_profiles.items():
        if len(profile_ids) > 1:
            errors.append(
                f"Product_Profiles: {set_id} product type {product_type!r} maps to "
                "multiple profiles"
            )

    accessory_profile_ids = {
        profile.profile_id for profile in profiles if profile.attribute_set_id == "mens_accessories"
    }
    missing_accessory_profiles = sorted(set(MANDATORY_ACCESSORY_PROFILES) - accessory_profile_ids)
    if "mens_accessories" in mappings and missing_accessory_profiles:
        errors.append(
            "Product_Profiles: mens_accessories is missing mandatory profiles "
            f"{missing_accessory_profiles}"
        )
    return errors


def _configuration_issues(
    mappings_by_set: dict[str, tuple[str, ...]],
    profiles: list[ProductProfile],
) -> dict[str, tuple[str, ...]]:
    rows_by_set = _group_rows(profiles, lambda row: row.attribute_set_id)
    issues: dict[str, tuple[str, ...]] = {}
    for set_id in mappings_by_set:
        set_profiles = rows_by_set.get(set_id, [])
        messages = []
        if not set_profiles:
            messages.append("No technical product profile is configured.")
        elif not any(profile.product_type for profile in set_profiles):
            messages.append("Approved CMS product types are absent.")
        if set_id != "topwear":
            messages.append("Approved set-specific permitted-value sources are absent.")
        issues[set_id] = tuple(messages)
    return issues


def load_registry(path: str | Path) -> Registry:
    path = Path(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise RegistryValidationError([f"cannot open {path}: {exc}"]) from exc

    errors: list[str] = []
    try:
        missing_sheets = [sheet for sheet in REQUIRED_COLUMNS if sheet not in workbook.sheetnames]
        if missing_sheets:
            raise RegistryValidationError([f"missing sheets {', '.join(missing_sheets)}"])

        tables = {sheet: _table_rows(workbook[sheet], sheet, errors) for sheet in REQUIRED_COLUMNS}
        attribute_sets = _parse_rows(
            AttributeSetRow, tables["Attribute_Sets"], "Attribute_Sets", errors
        )
        definitions = _parse_rows(
            AttributeDefinition,
            tables["Attribute_Definitions"],
            "Attribute_Definitions",
            errors,
        )
        permitted_headers = [
            str(value).strip()
            for value in next(
                workbook["Permitted_Values"].iter_rows(values_only=True),
                (),
            )
        ]
        value_columns = [header for header in permitted_headers if header.startswith("value_")]
        permitted_values = _parse_permitted_values(
            tables["Permitted_Values"], value_columns, errors
        )
        aliases = _parse_rows(ValueAlias, tables["Value_Aliases"], "Value_Aliases", errors)
        profiles = _parse_rows(
            ProductProfile, tables["Product_Profiles"], "Product_Profiles", errors
        )
    finally:
        workbook.close()

    if errors:
        raise RegistryValidationError(errors)
    errors = _validate_registry(attribute_sets, definitions, permitted_values, aliases, profiles)
    if errors:
        raise RegistryValidationError(errors)

    mappings_by_set = {
        set_id: tuple(row.header for row in sorted(rows, key=lambda item: item.position))
        for set_id, rows in _group_rows(attribute_sets, lambda row: row.attribute_set_id).items()
    }
    permitted_values_by_header = {row.attribute_header: row.values for row in permitted_values}
    aliases_by_header: dict[str, dict[str, str]] = defaultdict(dict)
    for alias in aliases:
        if not alias.active:
            continue
        canonical_lookup = {
            normalize_value(value): value
            for value in permitted_values_by_header[alias.attribute_header]
        }
        aliases_by_header[alias.attribute_header][normalize_value(alias.alias)] = canonical_lookup[
            normalize_value(alias.canonical_value)
        ]
    profiles_by_id = _group_rows(profiles, lambda row: (row.attribute_set_id, row.profile_id))

    return Registry(
        attribute_sets=tuple(attribute_sets),
        definitions=tuple(definitions),
        permitted_values=tuple(permitted_values),
        aliases=tuple(aliases),
        profiles=tuple(profiles),
        mappings_by_set=mappings_by_set,
        definitions_by_header={row.header: row for row in definitions},
        permitted_values_by_header=permitted_values_by_header,
        aliases_by_header=dict(aliases_by_header),
        profiles_by_id={key: tuple(rows) for key, rows in profiles_by_id.items()},
        configuration_issues_by_set=_configuration_issues(mappings_by_set, profiles),
        fingerprint=hashlib.sha256(path.read_bytes()).hexdigest(),
    )


def profile_ids(registry: Registry, attribute_set_id: str) -> tuple[str, ...]:
    if attribute_set_id not in registry.mappings_by_set:
        raise ValueError(f"Unknown attribute set {attribute_set_id!r}.")
    return tuple(
        dict.fromkeys(
            profile.profile_id
            for profile in registry.profiles
            if profile.attribute_set_id == attribute_set_id
        )
    )


def applicable_profile_headers(
    registry: Registry, attribute_set_id: str, profile_id: str
) -> tuple[str, ...]:
    if attribute_set_id not in registry.mappings_by_set:
        raise ValueError(f"Unknown attribute set {attribute_set_id!r}.")
    rows = registry.profiles_by_id.get((attribute_set_id, profile_id))
    if not rows:
        raise ValueError(f"Unknown profile {attribute_set_id}/{profile_id}.")
    applicable = {row.header for row in rows if row.applicable}
    return tuple(
        header for header in registry.mappings_by_set[attribute_set_id] if header in applicable
    )


def configuration_issues(registry: Registry, attribute_set_id: str) -> tuple[str, ...]:
    try:
        return registry.configuration_issues_by_set[attribute_set_id]
    except KeyError as exc:
        raise ValueError(f"Unknown attribute set {attribute_set_id!r}.") from exc


def _group_rows(rows: list, key: Any) -> dict[Any, list]:
    grouped: dict[Any, list] = defaultdict(list)
    for row in rows:
        grouped[key(row)].append(row)
    return grouped


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate an attribute registry workbook.")
    parser.add_argument("path", type=Path)
    args = parser.parse_args()
    registry = load_registry(args.path)
    print(
        f"valid: {len(registry.mappings_by_set)} sets, "
        f"{len(registry.definitions_by_header)} definitions, {registry.fingerprint}"
    )


if __name__ == "__main__":
    main()
