from __future__ import annotations

from collections import Counter, defaultdict
from io import BytesIO
from pathlib import PurePosixPath
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from fashion_cms.models import (
    MAX_EXCEL_CELL_CHARACTERS,
    InputRow,
    Severity,
    ValidationIssue,
    WorkbookResult,
)


REQUIRED_COLUMNS = (
    "sku",
    "base_code",
    "attributes__lulu_ean",
    "attributes__shipping_weight",
    "model_code_input_data",
)
SYSTEM_COPY_FIELDS = {
    "sku": "sku",
    "base_code": "base_code",
    "attributes__lulu_ean": "attributes__lulu_ean",
    "attributes__shipping_weight": "attributes__shipping_weight",
}
MAX_WORKBOOK_BYTES = 25 * 1024 * 1024
MAX_WORKBOOK_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_WORKBOOK_MEMBERS = 2_000
MAX_WORKBOOK_ROWS = 100_000
MAX_WORKBOOK_COLUMNS = 500
MAX_ROW_ISSUES = 100
BLANK_FIELD_MESSAGES = {
    "base_code": "Base code is blank; affected SKUs use their own internal group.",
    "attributes__lulu_ean": "EAN is blank.",
    "attributes__shipping_weight": "Shipping weight is blank.",
    "model_code_input_data": "Model input data is blank.",
}


def _issue(
    severity: Severity, code: str, message: str, location: str | None = None
) -> ValidationIssue:
    return ValidationIssue(severity=severity, code=code, message=message, location=location)


def _preflight_xlsx(content: bytes) -> ValidationIssue | None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
    except (BadZipFile, OSError) as exc:
        return _issue(Severity.CRITICAL, "MALFORMED_WORKBOOK", f"Cannot open workbook: {exc}")

    if len(members) > MAX_WORKBOOK_MEMBERS:
        return _issue(
            Severity.CRITICAL,
            "WORKBOOK_TOO_COMPLEX",
            f"Workbook contains more than {MAX_WORKBOOK_MEMBERS:,} internal files.",
        )
    if sum(member.file_size for member in members) > MAX_WORKBOOK_UNCOMPRESSED_BYTES:
        return _issue(
            Severity.CRITICAL,
            "WORKBOOK_TOO_LARGE",
            f"Workbook expands beyond {MAX_WORKBOOK_UNCOMPRESSED_BYTES // 1024 // 1024} MB.",
        )
    for member in members:
        path = PurePosixPath(member.filename.replace("\\", "/"))
        if path.is_absolute() or ".." in path.parts:
            return _issue(
                Severity.CRITICAL,
                "UNSAFE_WORKBOOK",
                f"Workbook contains an unsafe internal path: {member.filename}",
            )
        if member.flag_bits & 1:
            return _issue(
                Severity.CRITICAL,
                "ENCRYPTED_WORKBOOK",
                "Encrypted workbooks are not supported.",
            )
        if path.name.casefold() == "vbaproject.bin":
            return _issue(
                Severity.CRITICAL,
                "MACRO_WORKBOOK",
                "Macro-enabled workbook content is not supported.",
            )
    return None


def _duplicate_issue(
    values_to_rows: dict[str, list[int]], severity: Severity, code: str, label: str
) -> ValidationIssue | None:
    duplicates = [(value, rows) for value, rows in values_to_rows.items() if len(rows) > 1]
    if not duplicates:
        return None
    examples = "; ".join(
        f"{(value[:77] + '...' if len(value) > 80 else value)!r} "
        f"(rows {', '.join(map(str, row_numbers[:5]))})"
        for value, row_numbers in duplicates[:5]
    )
    if len(duplicates) > 5:
        examples += f"; and {len(duplicates) - 5:,} more"
    action = (
        "Rows were retained for review." if severity == Severity.WARNING else "Remove duplicates."
    )
    return _issue(
        severity,
        code,
        f"{len(duplicates):,} duplicate {label} value(s): {examples}. {action}",
    )


def parse_input_workbook(content: bytes, filename: str = "input.xlsx") -> WorkbookResult:
    issues: list[ValidationIssue] = []
    if PurePosixPath(filename).suffix.casefold() != ".xlsx":
        return WorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "UNSUPPORTED_WORKBOOK_TYPE",
                    "Upload a genuine .xlsx workbook; .xls is not supported.",
                ),
            )
        )
    if not content:
        return WorkbookResult(
            issues=(_issue(Severity.CRITICAL, "EMPTY_WORKBOOK", "The workbook is empty."),)
        )
    if len(content) > MAX_WORKBOOK_BYTES:
        return WorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "WORKBOOK_TOO_LARGE",
                    f"Workbook exceeds {MAX_WORKBOOK_BYTES // 1024 // 1024} MB.",
                ),
            )
        )
    if preflight_issue := _preflight_xlsx(content):
        return WorkbookResult(issues=(preflight_issue,))

    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        return WorkbookResult(
            issues=(
                _issue(Severity.CRITICAL, "MALFORMED_WORKBOOK", f"Cannot open workbook: {exc}"),
            )
        )

    try:
        if not workbook.worksheets:
            return WorkbookResult(
                issues=(
                    _issue(Severity.CRITICAL, "MISSING_WORKSHEET", "Workbook has no worksheet."),
                )
            )
        worksheet = workbook.worksheets[0]
        worksheet.reset_dimensions()
        cells_by_row = worksheet.iter_rows()
        header_cells = next(cells_by_row, None)
        if not header_cells:
            return WorkbookResult(
                issues=(
                    _issue(
                        Severity.CRITICAL,
                        "MISSING_HEADER_ROW",
                        "First worksheet has no header row.",
                        worksheet.title,
                    ),
                )
            )
        if len(header_cells) > MAX_WORKBOOK_COLUMNS:
            return WorkbookResult(
                issues=(
                    _issue(
                        Severity.CRITICAL,
                        "WORKSHEET_TOO_LARGE",
                        f"First worksheet must not exceed {MAX_WORKBOOK_COLUMNS:,} columns.",
                        worksheet.title,
                    ),
                )
            )
        formula_headers = [cell.coordinate for cell in header_cells if cell.data_type == "f"]
        if formula_headers:
            issues.append(
                _issue(
                    Severity.CRITICAL,
                    "FORMULA_NOT_ALLOWED",
                    "Header formulas are not allowed.",
                    f"{worksheet.title}!{', '.join(formula_headers)}",
                )
            )

        headers = [cell.value if isinstance(cell.value, str) else None for cell in header_cells]
        present_headers = [header for header in headers if header is not None]
        duplicate_headers = sorted(
            header for header, count in Counter(present_headers).items() if count > 1
        )
        if duplicate_headers:
            issues.append(
                _issue(
                    Severity.CRITICAL,
                    "DUPLICATE_COLUMNS",
                    f"Duplicate columns: {', '.join(duplicate_headers)}.",
                    worksheet.title,
                )
            )
        missing = [column for column in REQUIRED_COLUMNS if column not in present_headers]
        if missing:
            issues.append(
                _issue(
                    Severity.CRITICAL,
                    "MISSING_COLUMNS",
                    f"Missing required columns: {', '.join(missing)}.",
                    worksheet.title,
                )
            )
        if any(issue.severity == Severity.CRITICAL for issue in issues):
            return WorkbookResult(issues=tuple(issues))

        column_indexes = {header: headers.index(header) for header in REQUIRED_COLUMNS}
        rows: list[InputRow] = []
        blank_rows: dict[str, list[int]] = defaultdict(list)
        row_issue_count = 0
        row_issues_omitted = False

        def add_row_issue(issue: ValidationIssue) -> None:
            nonlocal row_issue_count, row_issues_omitted
            if row_issue_count < MAX_ROW_ISSUES:
                issues.append(issue)
                row_issue_count += 1
            else:
                row_issues_omitted = True

        for row_number, cells in enumerate(cells_by_row, start=2):
            if row_number > MAX_WORKBOOK_ROWS:
                issues.append(
                    _issue(
                        Severity.CRITICAL,
                        "WORKSHEET_TOO_LARGE",
                        f"First worksheet must not exceed {MAX_WORKBOOK_ROWS:,} rows.",
                        worksheet.title,
                    )
                )
                break
            if len(cells) > MAX_WORKBOOK_COLUMNS:
                issues.append(
                    _issue(
                        Severity.CRITICAL,
                        "WORKSHEET_TOO_LARGE",
                        f"First worksheet must not exceed {MAX_WORKBOOK_COLUMNS:,} columns.",
                        f"{worksheet.title}!{row_number}",
                    )
                )
                break
            if not any(
                cell.value is not None
                and (not isinstance(cell.value, str) or bool(cell.value.strip()))
                for cell in cells
            ):
                continue
            formula_cells = [cell.coordinate for cell in cells if cell.data_type == "f"]
            if formula_cells:
                add_row_issue(
                    _issue(
                        Severity.CRITICAL,
                        "FORMULA_NOT_ALLOWED",
                        "Workbook formulas are not accepted; replace them with text or values.",
                        f"{worksheet.title}!{', '.join(formula_cells)}",
                    )
                )
            required_cells = {
                header: cells[index] if index < len(cells) else None
                for header, index in column_indexes.items()
            }
            if any(cell is not None and cell.data_type == "f" for cell in required_cells.values()):
                continue
            error_fields = [
                header
                for header, cell in required_cells.items()
                if cell is not None and cell.data_type == "e"
            ]
            if error_fields:
                for field in error_fields:
                    column = get_column_letter(column_indexes[field] + 1)
                    add_row_issue(
                        _issue(
                            Severity.CRITICAL,
                            "INVALID_CELL",
                            f"{field}: Excel error cells are not accepted.",
                            f"{worksheet.title}!{column}{row_number}",
                        )
                    )
                continue

            values = {
                header: cell.value if cell is not None else None
                for header, cell in required_cells.items()
            }
            try:
                row = InputRow.model_validate({"row_number": row_number, **values})
            except ValidationError as exc:
                for error in exc.errors(include_url=False):
                    field = str(error["loc"][-1])
                    message = str(error["msg"]).removeprefix("Value error, ")
                    column = get_column_letter(column_indexes.get(field, 0) + 1)
                    add_row_issue(
                        _issue(
                            Severity.CRITICAL,
                            "INVALID_CELL",
                            f"{field}: {message}.",
                            f"{worksheet.title}!{column}{row_number}",
                        )
                    )
                continue

            rows.append(row)
            for field in BLANK_FIELD_MESSAGES:
                if getattr(row, field) is None:
                    blank_rows[field].append(row_number)
        if row_issues_omitted:
            issues.append(
                _issue(
                    Severity.CRITICAL,
                    "ADDITIONAL_ROW_ERRORS",
                    f"More than {MAX_ROW_ISSUES:,} row errors were found; additional errors "
                    "were omitted.",
                    worksheet.title,
                )
            )
        for field, row_numbers in blank_rows.items():
            sample = ", ".join(map(str, row_numbers[:10]))
            more = f", and {len(row_numbers) - 10:,} more" if len(row_numbers) > 10 else ""
            issues.append(
                _issue(
                    Severity.WARNING,
                    f"BLANK_{field.upper()}",
                    f"{BLANK_FIELD_MESSAGES[field]} Affected rows: {sample}{more}.",
                    worksheet.title,
                )
            )
    except Exception as exc:
        return WorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "MALFORMED_WORKBOOK",
                    f"Cannot read first worksheet: {exc}",
                ),
            )
        )
    finally:
        workbook.close()

    if not rows:
        issues.append(
            _issue(Severity.CRITICAL, "NO_VALID_ROWS", "Workbook contains no valid data rows.")
        )

    values_to_rows: dict[str, dict[str, list[int]]] = {
        "sku": defaultdict(list),
        "attributes__lulu_ean": defaultdict(list),
    }
    for row in rows:
        values_to_rows["sku"][row.sku].append(row.row_number)
        if row.attributes__lulu_ean:
            values_to_rows["attributes__lulu_ean"][row.attributes__lulu_ean].append(row.row_number)
    if duplicate_sku := _duplicate_issue(
        values_to_rows["sku"], Severity.CRITICAL, "DUPLICATE_SKU", "SKU"
    ):
        issues.append(duplicate_sku)
    if duplicate_ean := _duplicate_issue(
        values_to_rows["attributes__lulu_ean"],
        Severity.WARNING,
        "DUPLICATE_EAN",
        "EAN",
    ):
        issues.append(duplicate_ean)
    return WorkbookResult(rows=tuple(rows), issues=tuple(issues))


def build_blank_cms_workbook(rows: tuple[InputRow, ...], headers: tuple[str, ...]) -> bytes:
    if (
        not headers
        or any(
            not isinstance(header, str) or not header or len(header) > MAX_EXCEL_CELL_CHARACTERS
            for header in headers
        )
        or len(headers) != len(set(headers))
    ):
        raise ValueError("CMS headers must be non-empty and unique")
    for row in rows:
        for header, field in SYSTEM_COPY_FIELDS.items():
            value = getattr(row, field)
            if isinstance(value, str) and len(value) > MAX_EXCEL_CELL_CHARACTERS:
                raise ValueError(f"{header} exceeds the Excel cell character limit")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CMS Upload"
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.data_type = "s"

    for output_row, row in enumerate(rows, start=2):
        for column, header in enumerate(headers, start=1):
            field = SYSTEM_COPY_FIELDS.get(header)
            if field is None:
                continue
            value = getattr(row, field)
            cell = worksheet.cell(row=output_row, column=column, value=value)
            if header in {"sku", "base_code", "attributes__lulu_ean"}:
                cell.number_format = "@"
            if isinstance(value, str):
                cell.data_type = "s"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
