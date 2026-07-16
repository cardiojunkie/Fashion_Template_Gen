from __future__ import annotations

import ipaddress
import os
import random
import socket
import threading
import time
from collections.abc import Callable, Iterable, Mapping, Sequence
from concurrent.futures import FIRST_COMPLETED, Future, ThreadPoolExecutor, wait
from dataclasses import dataclass
from datetime import UTC, datetime
from email.utils import parsedate_to_datetime
from io import BytesIO
from pathlib import PurePosixPath
from urllib.parse import urljoin, urlsplit
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import httpx
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from fashion_cms.config import ResourceLimits
from fashion_cms.excel_service import preflight_xlsx
from fashion_cms.image_service import MAX_IMAGE_FILES, standardize_pad_white
from fashion_cms.models import (
    DownloadedImage,
    DownloadReportRow,
    DownloadResult,
    ImageDownloadResult,
    ImageUrlRequest,
    Severity,
    UrlWorkbookResult,
    ValidationIssue,
)


MIB = 1024 * 1024
MAX_URLS = MAX_IMAGE_FILES
MAX_REDIRECTS = 5
MAX_OUTPUT_FILENAME_CHARACTERS = 255
MAX_REPORT_ISSUES = 200
MAX_TOTAL_OUTPUT_BYTES = 500 * MIB
SUPPORTED_CONTENT_TYPES = {"image/jpeg", "image/png", "image/webp"}
REDIRECT_STATUSES = {301, 302, 303, 307, 308}
RETRYABLE_STATUSES = {408, 425, 429, 500, 502, 503, 504}
BLOCKED_HOSTS = {
    "instance-data",
    "instance-data.ec2.internal",
    "ip6-localhost",
    "localhost",
    "localhost.localdomain",
    "metadata",
    "metadata.azure.internal",
    "metadata.google.internal",
}
IPV6_TRANSITION_NETWORKS = tuple(
    ipaddress.ip_network(network)
    for network in ("64:ff9b::/96", "64:ff9b:1::/48", "2001::/32", "2002::/16")
)
REPORT_HEADERS = (
    "SKU",
    "URL ordinal",
    "Source URL",
    "Result",
    "HTTP status",
    "Output filename",
    "Source dimensions",
    "Output dimensions",
    "Error message",
)

Resolver = Callable[[str, int], Iterable[str]]
IpAddress = ipaddress.IPv4Address | ipaddress.IPv6Address


@dataclass(frozen=True)
class DownloadSettings:
    total_concurrency: int = 8
    per_host_concurrency: int = 4
    connect_timeout_seconds: float = 10.0
    read_timeout_seconds: float = 30.0
    retry_count: int = 3
    max_response_bytes: int = 25 * MIB
    max_decoded_pixels: int = 50_000_000
    max_total_output_bytes: int = MAX_TOTAL_OUTPUT_BYTES
    max_redirects: int = MAX_REDIRECTS
    total_deadline_seconds: float = 120.0
    max_urls: int = MAX_URLS
    max_image_dimension: int = 20_000

    def __post_init__(self) -> None:
        limits = (
            ("total_concurrency", self.total_concurrency, 1, 64),
            ("per_host_concurrency", self.per_host_concurrency, 1, 64),
            ("retry_count", self.retry_count, 0, 10),
            ("max_response_bytes", self.max_response_bytes, 1, 100 * MIB),
            ("max_decoded_pixels", self.max_decoded_pixels, 1, 100_000_000),
            ("max_total_output_bytes", self.max_total_output_bytes, 1, 1024 * MIB),
            ("max_redirects", self.max_redirects, 0, 20),
            ("max_urls", self.max_urls, 1, 10_000),
            ("max_image_dimension", self.max_image_dimension, 1, 65_535),
        )
        for name, value, minimum, maximum in limits:
            if (
                not isinstance(value, int)
                or isinstance(value, bool)
                or not minimum <= value <= maximum
            ):
                raise ValueError(f"{name} must be between {minimum:,} and {maximum:,}.")
        for name, value in (
            ("connect_timeout_seconds", self.connect_timeout_seconds),
            ("read_timeout_seconds", self.read_timeout_seconds),
        ):
            if isinstance(value, bool) or not isinstance(value, (int, float)) or not 0 < value <= 300:
                raise ValueError(f"{name} must be greater than 0 and at most 300.")
        if (
            isinstance(self.total_deadline_seconds, bool)
            or not isinstance(self.total_deadline_seconds, (int, float))
            or not 0 < self.total_deadline_seconds <= 900
        ):
            raise ValueError("total_deadline_seconds must be greater than 0 and at most 900.")

    @classmethod
    def from_env(cls, environ: Mapping[str, str] | None = None) -> DownloadSettings:
        values = os.environ if environ is None else environ
        resources = ResourceLimits.from_env(values)

        def integer(name: str, default: int) -> int:
            raw = values.get(name)
            if raw is None or not raw.strip():
                return default
            try:
                return int(raw)
            except ValueError as exc:
                raise ValueError(f"{name} must be an integer.") from exc

        def number(name: str, default: float) -> float:
            raw = values.get(name)
            if raw is None or not raw.strip():
                return default
            try:
                return float(raw)
            except ValueError as exc:
                raise ValueError(f"{name} must be a number.") from exc

        return cls(
            total_concurrency=integer("FASHION_CMS_IMAGE_TOTAL_CONCURRENCY", 8),
            per_host_concurrency=integer("FASHION_CMS_IMAGE_PER_HOST_CONCURRENCY", 4),
            connect_timeout_seconds=number(
                "FASHION_CMS_IMAGE_CONNECT_TIMEOUT_SECONDS",
                resources.url_connect_timeout_seconds,
            ),
            read_timeout_seconds=number(
                "FASHION_CMS_IMAGE_READ_TIMEOUT_SECONDS", resources.url_read_timeout_seconds
            ),
            retry_count=integer("FASHION_CMS_IMAGE_RETRY_COUNT", resources.url_retries),
            max_response_bytes=integer(
                "FASHION_CMS_IMAGE_MAX_RESPONSE_BYTES", resources.url_response_bytes
            ),
            max_decoded_pixels=integer(
                "FASHION_CMS_IMAGE_MAX_DECODED_PIXELS", resources.image_pixels
            ),
            max_redirects=resources.url_redirects,
            total_deadline_seconds=number(
                "FASHION_CMS_URL_TOTAL_DEADLINE_SECONDS",
                resources.url_total_deadline_seconds,
            ),
            max_total_output_bytes=resources.zip_expanded_bytes,
            max_urls=resources.uploaded_image_count,
            max_image_dimension=resources.image_dimension,
        )


class _DownloadFailure(Exception):
    def __init__(
        self,
        message: str,
        *,
        http_status: int | None = None,
        retryable: bool = False,
        retry_after: float | None = None,
    ) -> None:
        super().__init__(message)
        self.http_status = http_status
        self.retryable = retryable
        self.retry_after = retry_after


def _retry_after(value: str | None) -> float | None:
    if not value:
        return None
    try:
        return max(0.0, float(value))
    except ValueError:
        try:
            parsed = parsedate_to_datetime(value)
        except (TypeError, ValueError):
            return None
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=UTC)
        return max(0.0, (parsed - datetime.now(UTC)).total_seconds())


def _issue(
    severity: Severity,
    code: str,
    message: str,
    location: str | None = None,
) -> ValidationIssue:
    return ValidationIssue(severity=severity, code=code, message=message, location=location)


def _safe_output_filename(request: ImageUrlRequest) -> bool:
    filename = request.output_filename
    return (
        len(filename) <= MAX_OUTPUT_FILENAME_CHARACTERS
        and PurePosixPath(filename).name == filename
        and "\\" not in filename
        and not any(ord(character) < 32 or character in '<>:"|?*' for character in filename)
        and _xlsx_safe_text(filename) == filename
    )


def _xlsx_safe_text(value: str) -> str:
    return "".join(
        character
        if (
            character in "\t\n\r"
            or 0x20 <= ord(character) <= 0xD7FF
            or 0xE000 <= ord(character) <= 0xFFFD
            or 0x10000 <= ord(character) <= 0x10FFFF
        )
        and ord(character) not in {0xFFFE, 0xFFFF}
        else "�"
        for character in value
    )


def parse_url_workbook(content: bytes, filename: str = "image_urls.xlsx") -> UrlWorkbookResult:
    issues: list[ValidationIssue] = []
    resource_limits = ResourceLimits.from_env()
    urls_per_sku_limit = resource_limits.urls_per_sku
    url_limit = resource_limits.uploaded_image_count

    def add_issue(issue: ValidationIssue) -> None:
        if len(issues) < MAX_REPORT_ISSUES - 1:
            issues.append(issue)
        elif len(issues) == MAX_REPORT_ISSUES - 1:
            issues.append(
                _issue(
                    Severity.CRITICAL,
                    "ADDITIONAL_URL_ERRORS",
                    f"More than {MAX_REPORT_ISSUES:,} URL-workbook findings were detected; "
                    "additional findings were omitted.",
                )
            )

    if PurePosixPath(filename).suffix.casefold() != ".xlsx":
        return UrlWorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "UNSUPPORTED_WORKBOOK_TYPE",
                    "Upload a genuine .xlsx workbook; .xls is not supported.",
                ),
            )
        )
    if not isinstance(content, bytes) or not content:
        return UrlWorkbookResult(
            issues=(_issue(Severity.CRITICAL, "EMPTY_WORKBOOK", "The workbook is empty."),)
        )
    if len(content) > resource_limits.workbook_bytes:
        return UrlWorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "WORKBOOK_TOO_LARGE",
                    f"Workbook exceeds {resource_limits.workbook_bytes // MIB} MB.",
                ),
            )
        )
    if preflight_issue := preflight_xlsx(content, resource_limits):
        return UrlWorkbookResult(issues=(preflight_issue,))

    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
    except Exception:
        return UrlWorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "MALFORMED_WORKBOOK",
                    "Cannot open the URL workbook.",
                ),
            )
        )

    requests: list[ImageUrlRequest] = []
    output_keys: set[tuple[str, int]] = set()
    urls_by_sku: dict[str, int] = {}
    try:
        if not workbook.worksheets:
            add_issue(
                _issue(Severity.CRITICAL, "MISSING_WORKSHEET", "Workbook has no worksheet.")
            )
        else:
            worksheet = workbook.worksheets[0]
            worksheet.reset_dimensions()
            rows = worksheet.iter_rows()
            headers = next(rows, None)
            if not headers:
                add_issue(
                    _issue(
                        Severity.CRITICAL,
                        "MISSING_HEADER_ROW",
                        "First worksheet has no header row.",
                        worksheet.title,
                    )
                )
            elif len(headers) > resource_limits.workbook_columns:
                add_issue(
                    _issue(
                        Severity.CRITICAL,
                        "WORKSHEET_TOO_LARGE",
                        f"First worksheet must not exceed "
                        f"{resource_limits.workbook_columns:,} columns.",
                        worksheet.title,
                    )
                )
            else:
                formula_headers = [cell.coordinate for cell in headers if cell.data_type == "f"]
                if formula_headers:
                    add_issue(
                        _issue(
                            Severity.CRITICAL,
                            "FORMULA_NOT_ALLOWED",
                            "Header formulas are not allowed.",
                            f"{worksheet.title}!{', '.join(formula_headers)}",
                        )
                    )
                first_header = headers[0].value if headers else None
                if not isinstance(first_header, str) or first_header.strip().casefold() != "sku":
                    add_issue(
                        _issue(
                            Severity.CRITICAL,
                            "MISSING_SKU_COLUMN",
                            "Column A header must be SKU.",
                            worksheet.title,
                        )
                    )

                for row_number, cells in enumerate(rows, start=2):
                    if row_number > resource_limits.workbook_rows:
                        add_issue(
                            _issue(
                                Severity.CRITICAL,
                                "WORKSHEET_TOO_LARGE",
                                f"First worksheet must not exceed "
                                f"{resource_limits.workbook_rows:,} rows.",
                                worksheet.title,
                            )
                        )
                        break
                    if len(cells) > resource_limits.workbook_columns:
                        add_issue(
                            _issue(
                                Severity.CRITICAL,
                                "WORKSHEET_TOO_LARGE",
                                f"First worksheet must not exceed "
                                f"{resource_limits.workbook_columns:,} "
                                "columns.",
                                f"{worksheet.title}!{row_number}",
                            )
                        )
                        break
                    if not any(
                        cell.value is not None
                        and (not isinstance(cell.value, str) or bool(cell.value.strip()))
                        for cell in cells
                    ):
                        continue

                    sku_cell = cells[0]
                    url_cells = cells[1:]
                    populated_urls = [
                        cell
                        for cell in url_cells
                        if cell.value is not None
                        and (not isinstance(cell.value, str) or bool(cell.value.strip()))
                    ]
                    if not populated_urls:
                        continue
                    if sku_cell.data_type == "f":
                        add_issue(
                            _issue(
                                Severity.CRITICAL,
                                "FORMULA_NOT_ALLOWED",
                                "SKU formulas are not allowed.",
                                f"{worksheet.title}!{sku_cell.coordinate}",
                            )
                        )
                        continue
                    if sku_cell.data_type == "e" or not isinstance(sku_cell.value, str):
                        add_issue(
                            _issue(
                                Severity.CRITICAL,
                                "INVALID_SKU",
                                "SKU must be stored as text so leading zeros are preserved.",
                                f"{worksheet.title}!{sku_cell.coordinate}",
                            )
                        )
                        continue
                    sku = sku_cell.value.strip()
                    if not sku:
                        add_issue(
                            _issue(
                                Severity.CRITICAL,
                                "INVALID_SKU",
                                "SKU is required when a row contains image URLs.",
                                f"{worksheet.title}!{sku_cell.coordinate}",
                            )
                        )
                        continue
                    if len(sku_cell.value) > resource_limits.cell_characters:
                        add_issue(
                            _issue(
                                Severity.CRITICAL,
                                "INVALID_SKU",
                                "SKU exceeds the configured cell character limit.",
                                f"{worksheet.title}!{sku_cell.coordinate}",
                            )
                        )
                        continue

                    for ordinal, cell in enumerate(url_cells, start=1):
                        value = cell.value
                        if value is None or (isinstance(value, str) and not value.strip()):
                            continue
                        if cell.data_type == "f":
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "FORMULA_NOT_ALLOWED",
                                    "URL formulas are not allowed; provide literal URL text.",
                                    f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                            continue
                        if cell.data_type == "e" or not isinstance(value, str):
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "INVALID_URL_CELL",
                                    "Image URLs must be stored as text.",
                                    f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                            continue
                        if len(value) > resource_limits.cell_characters:
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "INVALID_URL_CELL",
                                    "Image URL exceeds the configured cell character limit.",
                                    f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                            continue
                        try:
                            request = ImageUrlRequest(
                                row_number=row_number,
                                sku=sku,
                                ordinal=ordinal,
                                source_url=value.strip(),
                            )
                        except ValidationError:
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "INVALID_URL_CELL",
                                    "SKU or image URL exceeds the supported Excel text limit.",
                                    f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                            continue
                        if not _safe_output_filename(request):
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "UNSAFE_OUTPUT_FILENAME",
                                    "SKU cannot be represented as a safe flat image filename.",
                                    f"{worksheet.title}!{sku_cell.coordinate}",
                                )
                            )
                            continue
                        output_key = (sku, ordinal)
                        if output_key in output_keys:
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "DUPLICATE_IMAGE_ORDINAL",
                                    f"SKU {sku[:80]!r} has more than one URL at ordinal "
                                    f"{ordinal}.",
                                    f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                            continue
                        output_keys.add(output_key)
                        urls_by_sku[sku] = urls_by_sku.get(sku, 0) + 1
                        if urls_by_sku[sku] > urls_per_sku_limit:
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "TOO_MANY_URLS_FOR_SKU",
                                    f"A SKU must not contain more than {urls_per_sku_limit} "
                                    "image URLs.",
                                    f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                            continue
                        requests.append(request)
                        if len(requests) > url_limit:
                            add_issue(
                                _issue(
                                    Severity.CRITICAL,
                                    "TOO_MANY_URLS",
                                    f"Workbook contains more than {url_limit:,} image URLs.",
                                    worksheet.title,
                                )
                            )
                            break
                    if len(requests) > url_limit:
                        break
    except Exception:
        return UrlWorkbookResult(
            issues=(
                _issue(
                    Severity.CRITICAL,
                    "MALFORMED_WORKBOOK",
                    "Cannot read the first worksheet.",
                ),
            )
        )
    finally:
        workbook.close()

    if len(requests) > url_limit:
        requests = requests[:url_limit]
    if not requests:
        add_issue(
            _issue(
                Severity.CRITICAL,
                "NO_IMAGE_URLS",
                "Workbook contains no usable image URLs.",
            )
        )
    return UrlWorkbookResult(requests=tuple(requests), issues=tuple(issues))


def resolve_host(host: str, port: int) -> tuple[str, ...]:
    return tuple(
        dict.fromkeys(
            result[4][0].split("%", 1)[0]
            for result in socket.getaddrinfo(
                host,
                port,
                family=socket.AF_UNSPEC,
                type=socket.SOCK_STREAM,
            )
        )
    )


def _is_public_address(address: IpAddress) -> bool:
    if (
        address.is_private
        or address.is_loopback
        or address.is_link_local
        or address.is_unspecified
        or address.is_multicast
        or address.is_reserved
        or not address.is_global
    ):
        return False
    if isinstance(address, ipaddress.IPv6Address):
        if address.is_site_local:
            return False
        if address.ipv4_mapped is not None and not address.ipv4_mapped.is_global:
            return False
        if any(address in network for network in IPV6_TRANSITION_NETWORKS):
            return False
    return True


def _validated_destination(url: str, resolver: Resolver) -> tuple[str, tuple[IpAddress, ...]]:
    if any(ord(character) < 32 or ord(character) == 127 for character in url):
        raise _DownloadFailure("URL contains invalid control characters.")
    try:
        parsed = urlsplit(url)
        port = parsed.port
    except (TypeError, ValueError) as exc:
        raise _DownloadFailure("URL is malformed.") from exc
    if parsed.scheme.casefold() not in {"http", "https"}:
        raise _DownloadFailure("Only HTTP and HTTPS image URLs are accepted.")
    if port is not None and port == 0:
        raise _DownloadFailure("URL port must be between 1 and 65535.")
    if parsed.username is not None or parsed.password is not None:
        raise _DownloadFailure("URLs containing credentials are not accepted.")
    if not parsed.hostname:
        raise _DownloadFailure("URL must include a hostname.")

    raw_host = parsed.hostname.rstrip(".").casefold()
    if not raw_host or any(ord(character) < 33 for character in raw_host):
        raise _DownloadFailure("URL hostname is invalid.")
    try:
        literal = ipaddress.ip_address(raw_host.split("%", 1)[0])
    except ValueError:
        try:
            host = raw_host.encode("idna").decode("ascii")
        except UnicodeError as exc:
            raise _DownloadFailure("URL hostname is invalid.") from exc
        if (
            host in BLOCKED_HOSTS
            or host.endswith(".localhost")
            or host.endswith(".local")
        ):
            raise _DownloadFailure("Local and metadata destinations are not allowed.")
        try:
            resolved = tuple(resolver(host, port or (443 if parsed.scheme == "https" else 80)))
        except (OSError, ValueError) as exc:
            raise _DownloadFailure("DNS lookup failed.", retryable=True) from exc
        if not resolved:
            raise _DownloadFailure("DNS lookup returned no addresses.", retryable=True)
        addresses: list[IpAddress] = []
        for value in resolved:
            try:
                addresses.append(ipaddress.ip_address(str(value).split("%", 1)[0]))
            except ValueError as exc:
                raise _DownloadFailure("DNS lookup returned an invalid address.") from exc
    else:
        host = raw_host
        addresses = [literal]

    if any(not _is_public_address(address) for address in addresses):
        raise _DownloadFailure("Private, local, and non-public destinations are not allowed.")
    return host, tuple(dict.fromkeys(addresses))


def _pinned_url(
    url: str,
    host: str,
    addresses: tuple[IpAddress, ...],
    offset: int,
) -> tuple[httpx.URL, dict[str, str], IpAddress]:
    parsed = urlsplit(url)
    address = addresses[offset % len(addresses)]
    authority = f"[{host}]" if ":" in host else host
    if parsed.port is not None:
        authority += f":{parsed.port}"
    return httpx.URL(url).copy_with(host=str(address)), {"Host": authority}, address


def _validate_connected_peer(
    response: httpx.Response,
    expected_address: IpAddress,
    *,
    required: bool,
) -> None:
    stream = response.extensions.get("network_stream")
    if stream is None or not hasattr(stream, "get_extra_info"):
        if required:
            raise _DownloadFailure("Connected server address could not be validated.")
        return
    try:
        server_address = stream.get_extra_info("server_addr")
        if not server_address:
            if required:
                raise _DownloadFailure("Connected server address could not be validated.")
            return
        peer = ipaddress.ip_address(str(server_address[0]).split("%", 1)[0])
    except (IndexError, TypeError, ValueError):
        raise _DownloadFailure("Connected server address could not be validated.") from None
    peer_v4 = peer.ipv4_mapped if isinstance(peer, ipaddress.IPv6Address) else None
    expected_v4 = (
        expected_address.ipv4_mapped
        if isinstance(expected_address, ipaddress.IPv6Address)
        else None
    )
    if not _is_public_address(peer) or not (
        peer == expected_address or peer_v4 == expected_address or expected_v4 == peer
    ):
        raise _DownloadFailure("Connected server address failed SSRF validation.")


def _host_semaphore(
    host: str,
    semaphores: dict[str, threading.BoundedSemaphore],
    lock: threading.Lock,
    per_host_limit: int,
) -> threading.BoundedSemaphore:
    with lock:
        return semaphores.setdefault(host, threading.BoundedSemaphore(per_host_limit))


def _fetch_once(
    request: ImageUrlRequest,
    *,
    client: httpx.Client,
    settings: DownloadSettings,
    resolver: Resolver,
    host_semaphores: dict[str, threading.BoundedSemaphore],
    host_lock: threading.Lock,
    address_offset: int,
    require_peer_address: bool,
    deadline: float,
    clock: Callable[[], float],
) -> tuple[bytes, int]:
    current_url = request.source_url
    destination = _validated_destination(current_url, resolver)

    for redirect_count in range(settings.max_redirects + 1):
        if clock() >= deadline:
            raise _DownloadFailure("Total request deadline was exceeded.")
        host, addresses = destination
        request_url, pinned_headers, expected_address = _pinned_url(
            current_url, host, addresses, address_offset
        )
        semaphore = _host_semaphore(
            host, host_semaphores, host_lock, settings.per_host_concurrency
        )
        try:
            remaining = deadline - clock()
            request_timeout = httpx.Timeout(
                connect=min(settings.connect_timeout_seconds, remaining),
                read=min(settings.read_timeout_seconds, remaining),
                write=min(settings.connect_timeout_seconds, remaining),
                pool=min(settings.connect_timeout_seconds, remaining),
            )
            with semaphore, client.stream(
                "GET",
                request_url,
                headers={
                    "Accept": "image/jpeg, image/png, image/webp",
                    "User-Agent": "Fashion-CMS-Image-Downloader/1",
                    **pinned_headers,
                },
                extensions={"sni_hostname": host}
                if request_url.scheme == "https"
                else None,
                timeout=request_timeout,
            ) as response:
                status = response.status_code
                _validate_connected_peer(
                    response,
                    expected_address,
                    required=require_peer_address,
                )
                if status in REDIRECT_STATUSES:
                    location = response.headers.get("location")
                    if not location:
                        raise _DownloadFailure(
                            "Redirect response did not include a destination.",
                            http_status=status,
                        )
                    if redirect_count == settings.max_redirects:
                        raise _DownloadFailure(
                            f"Redirect limit of {settings.max_redirects} was exceeded.",
                            http_status=status,
                        )
                    next_url = urljoin(current_url, location)
                    try:
                        destination = _validated_destination(next_url, resolver)
                    except _DownloadFailure as exc:
                        raise _DownloadFailure(
                            str(exc), http_status=status, retryable=exc.retryable
                        ) from exc
                    current_url = next_url
                    continue
                if not 200 <= status < 300:
                    raise _DownloadFailure(
                        f"HTTP {status} response.",
                        http_status=status,
                        retryable=status in RETRYABLE_STATUSES,
                        retry_after=_retry_after(response.headers.get("retry-after")),
                    )

                content_type = response.headers.get("content-type", "").split(";", 1)[0]
                if content_type.strip().casefold() not in SUPPORTED_CONTENT_TYPES:
                    raise _DownloadFailure(
                        "Response Content-Type is not a supported image type.",
                        http_status=status,
                    )
                content_length = response.headers.get("content-length")
                if content_length is not None:
                    try:
                        declared_size = int(content_length)
                    except ValueError as exc:
                        raise _DownloadFailure(
                            "Response Content-Length is invalid.", http_status=status
                        ) from exc
                    if declared_size < 0 or declared_size > settings.max_response_bytes:
                        raise _DownloadFailure(
                            f"Response exceeds the {settings.max_response_bytes // MIB} MB limit.",
                            http_status=status,
                        )

                body = bytearray()
                for chunk in response.iter_bytes(chunk_size=64 * 1024):
                    if clock() >= deadline:
                        raise _DownloadFailure(
                            "Total request deadline was exceeded.", http_status=status
                        )
                    if len(body) + len(chunk) > settings.max_response_bytes:
                        raise _DownloadFailure(
                            f"Response exceeds the {settings.max_response_bytes // MIB} MB limit.",
                            http_status=status,
                        )
                    body.extend(chunk)
                return bytes(body), status
        except _DownloadFailure:
            raise
        except httpx.TimeoutException as exc:
            raise _DownloadFailure("Request timed out.", retryable=True) from exc
        except httpx.TransportError as exc:
            raise _DownloadFailure("Network request failed.", retryable=True) from exc
        except httpx.StreamError as exc:
            raise _DownloadFailure("Network response stream failed.", retryable=True) from exc
        except httpx.InvalidURL as exc:
            raise _DownloadFailure("URL is malformed.") from exc

    raise _DownloadFailure("Redirect limit was exceeded.")


def _download_one(
    request: ImageUrlRequest,
    *,
    client: httpx.Client,
    settings: DownloadSettings,
    resolver: Resolver,
    sleeper: Callable[[float], None],
    jitter: Callable[[], float],
    host_semaphores: dict[str, threading.BoundedSemaphore],
    host_lock: threading.Lock,
    decode_semaphore: threading.BoundedSemaphore,
    require_peer_address: bool,
    clock: Callable[[], float],
) -> tuple[DownloadedImage | None, DownloadReportRow]:
    if not _safe_output_filename(request):
        return None, DownloadReportRow(
            sku=request.sku,
            ordinal=request.ordinal,
            source_url=request.source_url,
            result=DownloadResult.FAILED,
            error_message="SKU cannot be represented as a safe flat image filename.",
        )

    deadline = clock() + settings.total_deadline_seconds
    for attempt in range(settings.retry_count + 1):
        if clock() >= deadline:
            return None, DownloadReportRow(
                sku=request.sku,
                ordinal=request.ordinal,
                source_url=request.source_url,
                result=DownloadResult.FAILED,
                error_message="Total request deadline was exceeded.",
            )
        try:
            content, status = _fetch_once(
                request,
                client=client,
                settings=settings,
                resolver=resolver,
                host_semaphores=host_semaphores,
                host_lock=host_lock,
                address_offset=attempt,
                require_peer_address=require_peer_address,
                deadline=deadline,
                clock=clock,
            )
        except _DownloadFailure as exc:
            if exc.retryable and attempt < settings.retry_count:
                base_delay = min(8.0, 0.5 * 2**attempt)
                delay = (
                    exc.retry_after
                    if exc.retry_after is not None
                    else base_delay * (0.5 + 0.5 * min(max(jitter(), 0.0), 1.0))
                )
                remaining = deadline - clock()
                if remaining <= 0:
                    continue
                sleeper(min(delay, remaining))
                continue
            return None, DownloadReportRow(
                sku=request.sku,
                ordinal=request.ordinal,
                source_url=request.source_url,
                result=DownloadResult.FAILED,
                http_status=exc.http_status,
                error_message=str(exc),
            )

        try:
            # ponytail: serialize 50 MP decodes; raise this only with a measured memory budget.
            with decode_semaphore:
                standardized = standardize_pad_white(
                    content,
                    max_pixels=settings.max_decoded_pixels,
                    max_dimension=settings.max_image_dimension,
                )
        except ValueError as exc:
            return None, DownloadReportRow(
                sku=request.sku,
                ordinal=request.ordinal,
                source_url=request.source_url,
                result=DownloadResult.FAILED,
                http_status=status,
                error_message=str(exc),
            )

        image = DownloadedImage(
            sku=request.sku,
            ordinal=request.ordinal,
            source_url=request.source_url,
            output_filename=request.output_filename,
            source_width=standardized.source_dimensions[0],
            source_height=standardized.source_dimensions[1],
            output_width=standardized.output_dimensions[0],
            output_height=standardized.output_dimensions[1],
            low_resolution=standardized.low_resolution,
            content=standardized.content,
        )
        return image, DownloadReportRow(
            sku=request.sku,
            ordinal=request.ordinal,
            source_url=request.source_url,
            result=DownloadResult.SUCCESS,
            http_status=status,
            output_filename=image.output_filename,
            source_dimensions=standardized.source_dimensions,
            output_dimensions=standardized.output_dimensions,
        )

    raise AssertionError("retry loop must return")


def download_images(
    requests: Sequence[ImageUrlRequest],
    *,
    settings: DownloadSettings | None = None,
    transport: httpx.BaseTransport | None = None,
    resolver: Resolver = resolve_host,
    sleeper: Callable[[float], None] = time.sleep,
    jitter: Callable[[], float] = random.random,
    previous: ImageDownloadResult | None = None,
    clock: Callable[[], float] = time.monotonic,
) -> ImageDownloadResult:
    configuration = settings or DownloadSettings.from_env()
    request_list = tuple(requests)
    if len(request_list) > configuration.max_urls:
        raise ValueError(
            f"No more than {configuration.max_urls:,} image URLs may be downloaded at once."
        )
    if len({(request.sku, request.ordinal) for request in request_list}) != len(request_list):
        raise ValueError("Each SKU and URL ordinal must be unique.")

    previous_images = {image.key: image for image in previous.images} if previous else {}
    reused = {
        request.key: previous_images[request.key]
        for request in request_list
        if request.key in previous_images
    }
    pending = iter(request for request in request_list if request.key not in reused)
    completed_images = dict(reused)
    completed_reports: dict[tuple[str, int, str], DownloadReportRow] = {}
    if previous:
        prior_reports = {row.key: row for row in previous.report}
        for key, image in reused.items():
            completed_reports[key] = prior_reports.get(
                key,
                DownloadReportRow(
                    sku=image.sku,
                    ordinal=image.ordinal,
                    source_url=image.source_url,
                    result=DownloadResult.SUCCESS,
                    output_filename=image.output_filename,
                    source_dimensions=(image.source_width, image.source_height),
                    output_dimensions=(image.output_width, image.output_height),
                ),
            )

    timeout = httpx.Timeout(
        connect=configuration.connect_timeout_seconds,
        read=configuration.read_timeout_seconds,
        write=configuration.connect_timeout_seconds,
        pool=configuration.connect_timeout_seconds,
    )
    limits = httpx.Limits(
        max_connections=configuration.total_concurrency,
        max_keepalive_connections=0,
    )
    host_semaphores: dict[str, threading.BoundedSemaphore] = {}
    host_lock = threading.Lock()
    decode_semaphore = threading.BoundedSemaphore(1)
    retained_bytes = sum(len(image.content) for image in reused.values())

    with httpx.Client(
        timeout=timeout,
        follow_redirects=False,
        trust_env=False,
        transport=transport,
        limits=limits,
    ) as client, ThreadPoolExecutor(
        max_workers=configuration.total_concurrency,
        thread_name_prefix="image-download",
    ) as executor:
        active: dict[
            Future[tuple[DownloadedImage | None, DownloadReportRow]], ImageUrlRequest
        ] = {}

        def submit_next() -> bool:
            try:
                request = next(pending)
            except StopIteration:
                return False
            active[
                executor.submit(
                    _download_one,
                    request,
                    client=client,
                    settings=configuration,
                    resolver=resolver,
                    sleeper=sleeper,
                    jitter=jitter,
                    host_semaphores=host_semaphores,
                    host_lock=host_lock,
                    decode_semaphore=decode_semaphore,
                    require_peer_address=not isinstance(transport, httpx.MockTransport),
                    clock=clock,
                )
            ] = request
            return True

        for _ in range(configuration.total_concurrency):
            if not submit_next():
                break
        while active:
            done, _ = wait(active, return_when=FIRST_COMPLETED)
            for future in done:
                request = active.pop(future)
                try:
                    image, report = future.result()
                except Exception:
                    image = None
                    report = DownloadReportRow(
                        sku=request.sku,
                        ordinal=request.ordinal,
                        source_url=request.source_url,
                        result=DownloadResult.FAILED,
                        error_message="Image download failed unexpectedly.",
                    )
                if image and retained_bytes + len(image.content) > configuration.max_total_output_bytes:
                    image = None
                    report = DownloadReportRow(
                        sku=request.sku,
                        ordinal=request.ordinal,
                        source_url=request.source_url,
                        result=DownloadResult.FAILED,
                        http_status=report.http_status,
                        error_message="Processed images exceed the total output-size limit.",
                    )
                if image:
                    completed_images[request.key] = image
                    retained_bytes += len(image.content)
                completed_reports[request.key] = report
                submit_next()

    ordered_requests = sorted(request_list, key=lambda item: (item.sku, item.ordinal, item.source_url))
    return ImageDownloadResult(
        images=tuple(
            completed_images[request.key]
            for request in ordered_requests
            if request.key in completed_images
        ),
        report=tuple(completed_reports[request.key] for request in ordered_requests),
    )


def build_image_zip(images: Sequence[DownloadedImage]) -> bytes:
    ordered = sorted(images, key=lambda image: (image.sku, image.ordinal, image.source_url))
    names = [image.output_filename for image in ordered]
    if len(names) != len(set(names)) or any(PurePosixPath(name).name != name for name in names):
        raise ValueError("Image ZIP filenames must be unique and flat.")

    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
        for image in ordered:
            info = ZipInfo(image.output_filename, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            archive.writestr(info, image.content, compress_type=ZIP_DEFLATED, compresslevel=9)
    return output.getvalue()


def build_download_report(rows: Sequence[DownloadReportRow]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Image Download Report"
    for column, header in enumerate(REPORT_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.data_type = "s"

    ordered = sorted(rows, key=lambda row: (row.sku, row.ordinal, row.source_url))
    for row_number, row in enumerate(ordered, start=2):
        values: tuple[object, ...] = (
            row.sku,
            row.ordinal,
            row.source_url,
            row.result.value,
            row.http_status,
            row.output_filename,
            " × ".join(map(str, row.source_dimensions)) if row.source_dimensions else None,
            " × ".join(map(str, row.output_dimensions)) if row.output_dimensions else None,
            row.error_message,
        )
        for column, value in enumerate(values, start=1):
            if isinstance(value, str):
                value = _xlsx_safe_text(value)
            cell = worksheet.cell(row=row_number, column=column, value=value)
            if isinstance(value, str):
                cell.data_type = "s"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
