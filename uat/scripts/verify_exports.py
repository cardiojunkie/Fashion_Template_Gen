#!/usr/bin/env python3
"""Validate CMS exports without changing them."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import UTC, datetime
import json
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

from fashion_cms.excel_service import parse_input_workbook, preflight_xlsx
from fashion_cms.registry import (
    DataType,
    applicable_profile_headers,
    load_registry,
    profile_ids,
)


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_REGISTRY = ROOT / "config" / "attribute_registry.xlsx"
SYSTEM_HEADERS = {
    "sku",
    "base_code",
    "attributes__lulu_ean",
    "attributes__shipping_weight",
}
FORMULA_PREFIXES = ("=", "+", "-", "@")
INTERNAL_MARKERS = (
    "canonical_value",
    "confidence",
    "evidence",
    "normalization",
    "prompt_version",
    "raw_value",
    "request_id",
    "review_status",
    "schema_version",
)


def _safe_expected(value: str | None) -> str | None:
    if value and value.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _finding(code: str, message: str, location: str | None = None) -> dict[str, str]:
    finding = {"code": code, "message": message}
    if location:
        finding["location"] = location
    return finding


def _expected_rows(input_workbook: Path) -> tuple[dict[str, dict[str, str | None]], list[dict]]:
    parsed = parse_input_workbook(input_workbook.read_bytes(), input_workbook.name)
    errors = [
        _finding(issue.code, issue.message, issue.location)
        for issue in parsed.issues
        if issue.severity.value == "CRITICAL"
    ]
    return (
        {
            row.sku: {
                "sku": _safe_expected(row.sku),
                "base_code": _safe_expected(row.base_code),
                "attributes__lulu_ean": _safe_expected(row.attributes__lulu_ean),
            }
            for row in parsed.rows
        },
        errors,
    )


def verify_export(
    workbook_path: str | Path,
    *,
    attribute_set: str,
    input_workbook: str | Path,
    registry_path: str | Path = DEFAULT_REGISTRY,
    profile: str | None = None,
) -> dict[str, Any]:
    """Return a JSON-safe verification result for one exported workbook."""
    path = Path(workbook_path)
    registry = load_registry(registry_path)
    if attribute_set not in registry.mappings_by_set:
        raise ValueError(f"Unknown attribute set {attribute_set!r}.")
    available_profiles = profile_ids(registry, attribute_set)
    if profile is None and len(available_profiles) == 1:
        profile = available_profiles[0]
    if attribute_set == "mens_accessories" and profile is None:
        raise ValueError("--profile is required for Men's Accessories exports.")
    if profile is not None and profile not in available_profiles:
        raise ValueError(
            f"Unknown profile {attribute_set}/{profile}; choose from "
            f"{', '.join(available_profiles)}."
        )

    expected_rows, input_errors = _expected_rows(Path(input_workbook))
    findings = list(input_errors)
    result: dict[str, Any] = {
        "workbook": str(path),
        "attribute_set": attribute_set,
        "profile": profile,
        "status": "FAIL",
        "header_count": None,
        "row_count": None,
        "findings": findings,
    }
    if input_errors:
        findings.append(
            _finding("INVALID_INPUT_WORKBOOK", "Expected-SKU workbook has critical errors.")
        )
        return result

    try:
        preflight_issue = preflight_xlsx(path.read_bytes())
    except OSError as exc:
        findings.append(_finding("WORKBOOK_OPEN", f"Cannot read workbook: {exc}"))
        return result
    if preflight_issue:
        findings.append(
            _finding(preflight_issue.code, preflight_issue.message, preflight_issue.location)
        )
        return result

    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception as exc:
        findings.append(_finding("WORKBOOK_OPEN", f"Cannot open workbook: {exc}"))
        return result

    try:
        if len(workbook.worksheets) != 1:
            findings.append(
                _finding(
                    "WORKSHEET_COUNT",
                    f"Expected one CMS worksheet; found {len(workbook.worksheets)}.",
                )
            )
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows()
        header_cells = next(rows, ())
        actual_headers = tuple(cell.value for cell in header_cells)
        expected_headers = registry.mappings_by_set[attribute_set]
        result["header_count"] = len(actual_headers)
        if actual_headers != expected_headers:
            findings.append(
                _finding(
                    "HEADERS",
                    "CMS headers do not exactly match the canonical names, membership, and order.",
                    f"{worksheet.title}!1",
                )
            )
        extras = [header for header in actual_headers if header not in expected_headers]
        internal = [
            str(header)
            for header in extras
            if any(marker in str(header).casefold() for marker in INTERNAL_MARKERS)
        ]
        if internal:
            findings.append(
                _finding("INTERNAL_COLUMNS", f"Internal/debug columns found: {', '.join(internal)}.")
            )

        indexes = {header: index for index, header in enumerate(actual_headers)}
        data_rows = [row for row in rows if any(cell.value is not None for cell in row)]
        result["row_count"] = len(data_rows)
        sku_index = indexes.get("sku")
        if sku_index is None:
            findings.append(_finding("MISSING_SKU", "The CMS workbook has no sku column."))
            return result

        output_skus: list[str] = []
        rows_by_sku: dict[str, Sequence[Any]] = {}
        for row_number, row in enumerate(data_rows, start=2):
            sku_cell = row[sku_index] if sku_index < len(row) else None
            sku = sku_cell.value if sku_cell is not None else None
            if not isinstance(sku, str):
                findings.append(
                    _finding("SKU_TEXT", "SKU must be stored as text.", f"{worksheet.title}!{row_number}")
                )
                continue
            output_skus.append(sku)
            rows_by_sku[sku] = row

        duplicates = sorted(sku for sku, count in Counter(output_skus).items() if count > 1)
        if duplicates:
            findings.append(_finding("DUPLICATE_SKU", f"Duplicate output SKUs: {duplicates}."))
        expected_output_skus = {values["sku"] for values in expected_rows.values()}
        missing = sorted(expected_output_skus - set(output_skus))
        unexpected = sorted(set(output_skus) - expected_output_skus)
        if len(data_rows) != len(expected_rows) or missing or unexpected:
            findings.append(
                _finding(
                    "SKU_ROWS",
                    f"Expected {len(expected_rows)} rows; found {len(data_rows)}. "
                    f"Missing: {missing or 'none'}; unexpected: {unexpected or 'none'}.",
                )
            )

        for source_sku, expected in expected_rows.items():
            output_sku = expected["sku"]
            row = rows_by_sku.get(output_sku)
            if row is None:
                continue
            for header in ("sku", "base_code", "attributes__lulu_ean"):
                index = indexes.get(header)
                if index is None or index >= len(row):
                    continue
                cell = row[index]
                value = cell.value
                if value is not None and not isinstance(value, str):
                    findings.append(
                        _finding(
                            "IDENTIFIER_TEXT",
                            f"{header} for SKU {source_sku!r} is not stored as text.",
                            cell.coordinate,
                        )
                    )
                if value != expected[header]:
                    findings.append(
                        _finding(
                            "IDENTIFIER_VALUE",
                            f"{header} for SKU {source_sku!r} does not preserve the input text.",
                            cell.coordinate,
                        )
                    )

        for row in data_rows:
            for cell in row:
                value = cell.value
                if cell.data_type == "f" or (
                    isinstance(value, str) and value.startswith(FORMULA_PREFIXES)
                ):
                    findings.append(
                        _finding(
                            "UNSAFE_FORMULA",
                            "Formula or formula-like text is not safely neutralized.",
                            cell.coordinate,
                        )
                    )

        for header in expected_headers:
            definition = registry.definitions_by_header[header]
            if definition.data_type != DataType.ENUM or header not in indexes:
                continue
            permitted = set(registry.permitted_values_by_header[header])
            index = indexes[header]
            for row in data_rows:
                if index >= len(row) or row[index].value in (None, ""):
                    continue
                if row[index].value not in permitted:
                    findings.append(
                        _finding(
                            "INVALID_ENUM",
                            f"{header} contains non-permitted value {row[index].value!r}.",
                            row[index].coordinate,
                        )
                    )

        if profile is not None:
            allowed = set(applicable_profile_headers(registry, attribute_set, profile))
            allowed.update(SYSTEM_HEADERS)
            allowed.update(
                header
                for header in expected_headers
                if registry.definitions_by_header[header].data_type == DataType.GENERATED_TEXT
            )
            for header in set(expected_headers) - allowed:
                index = indexes.get(header)
                if index is None:
                    continue
                for row in data_rows:
                    if index < len(row) and row[index].value not in (None, ""):
                        findings.append(
                            _finding(
                                "PROFILE_INAPPLICABLE",
                                f"{header} must be blank for profile {profile}.",
                                row[index].coordinate,
                            )
                        )
        result["status"] = "PASS" if not findings else "FAIL"
        return result
    finally:
        workbook.close()


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Validate exact CMS exports without modifying the workbooks."
    )
    parser.add_argument("workbooks", nargs="+", type=Path, help="CMS .xlsx exports to verify")
    parser.add_argument("--attribute-set", required=True, help="Canonical attribute-set ID")
    parser.add_argument(
        "--input-workbook",
        required=True,
        type=Path,
        help="Original input workbook used to establish expected SKU rows and identifiers",
    )
    parser.add_argument(
        "--profile",
        help="Product profile; required for Men's Accessories and optional otherwise",
    )
    parser.add_argument("--registry", type=Path, default=DEFAULT_REGISTRY)
    parser.add_argument(
        "--report-json",
        type=Path,
        default=Path("uat_export_report.json"),
        help="Machine-readable output path (default: ./uat_export_report.json)",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        results = [
            verify_export(
                path,
                attribute_set=args.attribute_set,
                input_workbook=args.input_workbook,
                registry_path=args.registry,
                profile=args.profile,
            )
            for path in args.workbooks
        ]
    except (OSError, ValueError) as exc:
        _parser().error(str(exc))
    report = {
        "report_version": "1",
        "generated_at": datetime.now(UTC).isoformat(),
        "status": "PASS" if all(result["status"] == "PASS" for result in results) else "FAIL",
        "results": results,
    }
    args.report_json.parent.mkdir(parents=True, exist_ok=True)
    args.report_json.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    for result in results:
        print(
            f"{result['status']}: {result['workbook']} "
            f"({result['header_count']} headers, {result['row_count']} rows)"
        )
        for finding in result["findings"]:
            location = f" [{finding['location']}]" if finding.get("location") else ""
            print(f"  - {finding['code']}{location}: {finding['message']}")
    print(f"Machine-readable report: {args.report_json}")
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
