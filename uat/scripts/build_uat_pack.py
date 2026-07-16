#!/usr/bin/env python3
"""Regenerate the manual-UAT workbooks and header contract."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image

from fashion_cms.excel_service import REQUIRED_COLUMNS
from fashion_cms.registry import load_registry


ROOT = Path(__file__).resolve().parents[2]
UAT = ROOT / "uat"
INPUTS = UAT / "inputs"
REGISTRY_PATH = ROOT / "config" / "attribute_registry.xlsx"
CHECKLIST_COLUMNS = (
    "Test ID",
    "Module",
    "Preconditions",
    "Input/Test Data",
    "Steps",
    "Expected Result",
    "Actual Result",
    "Status",
    "Evidence/Screenshot",
    "Defect ID",
    "Notes",
)
STATUSES = "NOT STARTED,PASS,FAIL,BLOCKED,NOT APPLICABLE"
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
PASS_FILL = PatternFill("solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")


def _fit_sheet(sheet, widths: dict[str, int] | None = None) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, cell in enumerate(sheet[1], start=1):
        width = (widths or {}).get(str(cell.value), 18)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _checklist_sheet(workbook: Workbook, title: str, rows: list[tuple[str, ...]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(CHECKLIST_COLUMNS)
    for row in rows:
        sheet.append(row)
    status_column = CHECKLIST_COLUMNS.index("Status") + 1
    validation = DataValidation(type="list", formula1=f'"{STATUSES}"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add(f"{get_column_letter(status_column)}2:{get_column_letter(status_column)}1000")
    sheet.conditional_formatting.add(
        f"{get_column_letter(status_column)}2:{get_column_letter(status_column)}1000",
        FormulaRule(formula=[f'${get_column_letter(status_column)}2="PASS"'], fill=PASS_FILL),
    )
    sheet.conditional_formatting.add(
        f"{get_column_letter(status_column)}2:{get_column_letter(status_column)}1000",
        FormulaRule(formula=[f'${get_column_letter(status_column)}2="FAIL"'], fill=FAIL_FILL),
    )
    _fit_sheet(
        sheet,
        {
            "Test ID": 16,
            "Module": 22,
            "Preconditions": 28,
            "Input/Test Data": 34,
            "Steps": 52,
            "Expected Result": 52,
            "Actual Result": 30,
            "Status": 18,
            "Evidence/Screenshot": 28,
            "Defect ID": 14,
            "Notes": 30,
        },
    )


def _row(
    test_id: str,
    module: str,
    preconditions: str,
    data: str,
    steps: str,
    expected: str,
) -> tuple[str, ...]:
    return (test_id, module, preconditions, data, steps, expected, "", "NOT STARTED", "", "", "")


def build_checklist() -> None:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(("Topic", "Guidance"))
    for topic, guidance in (
        ("Purpose", "Record manual user acceptance. Automated evidence is not user approval."),
        ("How to run", "Follow uat/README.md in order. Enter Actual Result and attach evidence."),
        ("Status", "Choose NOT STARTED, PASS, FAIL, BLOCKED, or NOT APPLICABLE."),
        ("Failures", "Create a Defect ID, record exact steps and evidence, and do not sign off."),
        ("Live vision", "Use user-supplied real products. If no approved API setup exists, mark BLOCKED."),
        ("Production", "Passing this workbook does not deploy or approve production automatically."),
    ):
        instructions.append((topic, guidance))
    _fit_sheet(instructions, {"Topic": 20, "Guidance": 100})

    environment = workbook.create_sheet("Environment")
    environment.append(("Field", "Value", "Notes"))
    for field in (
        "Tester name",
        "Test date/time UTC",
        "Repository commit",
        "Release candidate",
        "Codespace/browser",
        "Python version",
        "Registry fingerprint",
        "Extraction mode/model",
        "Port visibility (must be private)",
        "Ground-truth workbook path",
    ):
        environment.append((field, "", ""))
    _fit_sheet(environment, {"Field": 34, "Value": 44, "Notes": 60})

    _checklist_sheet(
        workbook,
        "Phase Audit",
        [
            _row(
                f"PHASE-{phase}",
                f"Phase {phase}",
                "Automated suite and registry validation passed.",
                "docs/RETROSPECTIVE_AUDIT.md",
                "Open the audit matrix; sample the cited implementation and evidence for this phase.",
                "Statuses are evidence-backed; blocked user decisions are not marked PASS.",
            )
            for phase in range(1, 9)
        ],
    )
    _checklist_sheet(
        workbook,
        "CMS Generator",
        [
            _row("CMS-001", "Startup", "App is running privately.", "Home page", "Open the forwarded URL.", "CMS Generator loads without an exception."),
            _row("CMS-002", "Valid workbook", "Select any set/profile.", "matching *_structural.xlsx", "Upload workbook and matched images.", "Rows and identifiers preview; no critical error."),
            _row("CMS-003", "Duplicate SKU", "App open.", "duplicate_sku.xlsx", "Upload workbook.", "Exact duplicate SKU is reported and processing is blocked."),
            _row("CMS-004", "Required columns", "App open.", "missing_required_column.xlsx", "Upload workbook.", "Missing column is named and processing is blocked."),
            _row("CMS-005", "Identifiers", "Use Topwear structural input.", "leading-zero SKU/EAN row", "Preview, export blank CMS, reopen it.", "Leading zeros and base code remain text and unchanged."),
            _row("CMS-006", "Warnings", "App open.", "missing_base_code.xlsx", "Upload workbook.", "Blank base code warns; internal fallback never enters export."),
            _row("CMS-007", "Images", "Use uat/inputs/images.", "matched, missing, orphan, duplicate ordinal", "Upload cases separately.", "Exact files/SKUs/ordinals are identified before any model call."),
        ],
    )
    _checklist_sheet(
        workbook,
        "Variant Testing",
        [
            _row("VAR-001", "Default mode", "Valid multi-SKU workbook uploaded.", "Any base-code group", "Create job and inspect groups.", "Every group starts PER_SKU."),
            _row("VAR-002", "Safe size-only", "Use two user-verified genuine size variants.", "Same product; only size differs", "Select BASE_CODE_SIZE_ONLY and confirm; inspect planned calls and output.", "One call; shared eligible visual facts; SKU/EAN/size remain row-specific."),
            _row("VAR-003", "Unsafe visual variants", "Use real black solid and blue striped products with same base code.", "User-supplied images and facts", "Keep PER_SKU; inspect warnings, call count, review, and export.", "One call per SKU; no color/pattern/design leakage in either direction."),
            _row("VAR-004", "Mixed modes", "At least two base-code groups.", "One safe size-only and one varying group", "Choose a different mode per group.", "Mixed modes persist and planned request count is exact."),
            _row("VAR-005", "Representative", "Size-only group has unequal image counts.", "User-selected representative", "Change representative, save, restart app.", "Selection is editable and persists; default is deterministic."),
            _row("VAR-006", "Malformed size", "Structural workbook uploaded.", "model data containing XXL–L", "Inspect grouping and review.", "Malformed size is not silently normalized or used to assert safe grouping."),
        ],
    )
    _checklist_sheet(
        workbook,
        "Review and Catalog Copy",
        [
            _row("REV-001", "Source priority", "Fake or live extraction complete.", "Explicit input conflicting with visual proposal", "Open review.", "Explicit input wins; conflict and evidence remain visible."),
            _row("REV-002", "Actions", "Review item exists.", "One item per action", "Accept, edit, reject, and blank; restart app.", "Every decision persists with note/provenance."),
            _row("REV-003", "Enum safety", "Enum review item exists.", "Unsupported value", "Try to accept/edit invalid enum.", "Invalid enum cannot be silently accepted; cell stays blank/review-required."),
            _row("REV-004", "Vision color", "Color absent from input.", "Approved broad visual color", "Accept unchanged image suggestion.", "Suggestion stays review-visible and accepted cell is identified as image-derived."),
            _row("COPY-001", "Titles", "Review complete.", "Accepted facts with missing components", "Generate copy.", "Title/name use accepted facts only; missing parts omitted; model not duplicated."),
            _row("COPY-002", "Bullets", "Review complete.", "Sparse accepted facts", "Generate copy and inspect six cells.", "Factual bullets only; unsupported cells blank; openings are not repetitive."),
            _row("COPY-003", "Keywords", "Review complete.", "Accepted customer-facing facts", "Generate copy.", "Useful keywords without stuffing or unsupported claims."),
        ],
    )
    _checklist_sheet(
        workbook,
        "Attribute Set Exports",
        [
            _row(
                f"EXPORT-{index:02d}",
                name,
                "Complete fake-client workflow and review.",
                f"{set_id}_structural.xlsx",
                "Export CMS and QC separately; reopen CMS; run verify_exports.py.",
                f"Exact {count} headers/order, expected SKU rows, text identifiers, no internal fields.",
            )
            for index, (set_id, name, count) in enumerate(
                (
                    ("topwear", "Topwear", 45),
                    ("bottomwear", "Bottomwear", 43),
                    ("ethnic_wear", "Ethnic Wear", 44),
                    ("inner_sleepwear", "Innerwear & Sleepwear", 43),
                    ("footwear", "Footwear", 46),
                    ("sports_activewear", "Sports & Activewear", 46),
                    ("mens_accessories", "Men's Accessories", 61),
                ),
                start=1,
            )
        ],
    )
    _checklist_sheet(
        workbook,
        "Men’s Accessories",
        [
            _row(
                f"ACC-{index:02d}",
                profile,
                "Men's Accessories selected; use a real matching product.",
                "User-supplied product and approved ground truth",
                "Select and confirm profile; extract, review, export, and verify with --profile.",
                "Only profile-relevant specialist fields appear; all other specialist CMS cells are blank.",
            )
            for index, profile in enumerate(
                ("bags_luggage", "caps_headwear", "watches", "eyewear", "belts_wallets_ties_other"),
                start=1,
            )
        ],
    )
    _checklist_sheet(
        workbook,
        "Image Downloader",
        [
            _row("IMG-DL-001", "Valid image", "Enter a stable public test URL.", "image_downloader_uat.xlsx", "Download and inspect ZIP/report.", "1500×1500 RGB JPEG; white canvas; aspect ratio preserved; flat ZIP."),
            _row("IMG-DL-002", "Transparency", "Enter a public transparent PNG URL.", "Image 1", "Download and inspect corners/background.", "Transparent areas become white."),
            _row("IMG-DL-003", "Ordinal", "Enter failing URL 1 and valid URL 2.", "Images 1 and 2", "Download.", "Success is sku-2.jpg, never renamed sku-1.jpg."),
            _row("IMG-DL-004", "Malformed content", "Enter URL returning HTML.", "Image 3", "Download.", "HTML is rejected and detailed in report; absent from ZIP."),
            _row("IMG-DL-005", "SSRF", "No external request should be sent.", "http://127.0.0.1", "Enter URL and download.", "Local/private destination is blocked safely."),
            _row("IMG-DL-006", "Retry", "At least one success and failure.", "Prior batch", "Fix failed URL and choose Retry failed URLs.", "Successful URL is not downloaded again."),
        ],
    )
    _checklist_sheet(
        workbook,
        "Job Recovery",
        [
            _row("JOB-001", "Partial failure", "Multi-item fake job.", "One controlled fake failure", "Run job.", "Successful items remain; failed item is isolated."),
            _row("JOB-002", "Retry", "Partial failure exists.", "Failed item only", "Retry failed work.", "Successful calls are not repeated."),
            _row("JOB-003", "Restart", "Persisted job exists.", "Same validated inputs", "Stop/restart app and reopen Job History.", "Job state, group choices, errors, and reviews persist."),
            _row("JOB-004", "Cancellation", "Multi-item job running.", "Fake job", "Request cancellation, then resume.", "Completed results remain; unscheduled work resumes safely."),
            _row("JOB-005", "Partial export", "Some items succeeded.", "Partial job", "Generate/export successful work.", "CMS has successful rows only; QC lists incomplete SKUs."),
        ],
    )
    _checklist_sheet(
        workbook,
        "Security Checks",
        [
            _row("SEC-001", "Formula", "App open.", "formula_like_text.xlsx", "Upload, process if allowed, export, reopen.", "Text is rejected or safely neutralized; no formula executes."),
            _row("SEC-002", "Malformed image", "App open.", "images/malformed.jpg", "Upload image.", "Decode failure is actionable and blocks unsafe processing."),
            _row("SEC-003", "Private URL", "Image Downloader open.", "http://127.0.0.1", "Attempt download.", "Request is rejected before local access."),
            _row("SEC-004", "Path filename", "App open.", "path-like or ZIP traversal filename", "Upload benign crafted fixture.", "Unsafe path is rejected; no file is extracted outside the archive boundary."),
            _row("SEC-005", "Extension", "App open.", "unsupported_extension.txt", "Upload if browser allows, otherwise record browser block.", "Unsupported type is rejected or unavailable for selection."),
            _row("SEC-006", "Limits", "Use safe local test data near configured limits.", "Operator-created fixture", "Upload and observe.", "Limit is clear; app stays responsive; no data is silently truncated."),
        ],
    )
    _checklist_sheet(workbook, "Defects", [])
    _checklist_sheet(
        workbook,
        "User Sign-Off",
        [
            _row(
                f"SIGN-{index:02d}",
                "Business decision",
                "Authorized approver and source are available.",
                decision,
                "Record the exact approved rule/source/version in Notes and supporting repository sign-off document.",
                "Decision is explicit, versioned, attributable, and has a rollback instruction.",
            )
            for index, decision in enumerate(
                (
                    "Final permitted values",
                    "Semantic attribute pairs",
                    "Titles and character limits",
                    "Bullet rules",
                    ".xlsx/.xls requirement",
                    "Background removal requirement",
                    "Retention period",
                    "Production hosting",
                    "Authentication",
                    "Approved models",
                    "Approved pricing",
                    "Auto-accept thresholds",
                ),
                start=1,
            )
        ],
    )
    workbook.save(UAT / "manual_uat_checklist.xlsx")
    workbook.close()


def _write_input(path: Path, rows: list[tuple[object, ...]], columns=REQUIRED_COLUMNS) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet.append(columns)
    for row in rows:
        sheet.append(row)
    for column in ("A", "B", "C"):
        for cell in sheet[column][1:]:
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = "@"
    _fit_sheet(sheet, {column: 28 for column in columns})
    workbook.save(path)
    workbook.close()


def build_inputs() -> None:
    rows = [
        ("000123", "STRUCT-LEADING", "0000000123456", "0.25", "Structural identifier test only."),
        ("ABC-12", "STRUCT-HYPHEN", "0000000123457", "0.30", "Structural hyphenated SKU test only."),
        ("REPLACE-SIZE-1", "REPLACE-SIZE-GROUP", "0000000123458", "", ""),
        ("REPLACE-SIZE-2", "REPLACE-SIZE-GROUP", "0000000123459", "", ""),
        ("REPLACE-VISUAL-1", "REPLACE-VISUAL-GROUP", "0000000123460", "", ""),
        ("REPLACE-VISUAL-2", "REPLACE-VISUAL-GROUP", "0000000123461", "", ""),
        ("MALFORMED-SIZE", "STRUCT-MALFORMED", "0000000123462", "", "Intentional malformed size test: size=XXL–L."),
        ("MISSING-IMAGE", "STRUCT-MISSING-IMAGE", "0000000123463", "", "Structural missing-image test only."),
        ("FORMULA-TEXT", "STRUCT-FORMULA", "0000000123464", "", "+1+1 untrusted formula-like text test."),
    ]
    for set_id in load_registry(REGISTRY_PATH).mappings_by_set:
        _write_input(INPUTS / f"{set_id}_structural.xlsx", rows)
    _write_input(
        INPUTS / "duplicate_sku.xlsx",
        [
            ("DUPLICATE-1", "DUP", "0000000123500", "", "First duplicate row."),
            ("DUPLICATE-1", "DUP", "0000000123501", "", "Second duplicate row."),
        ],
    )
    _write_input(
        INPUTS / "missing_base_code.xlsx",
        [("BLANK-BASE", "", "0000000123502", "", "Base code intentionally blank.")],
    )
    _write_input(
        INPUTS / "formula_like_text.xlsx",
        [("FORMULA-LIKE", "FORMULA", "0000000123503", "", "+1+1 benign formula-like text.")],
    )
    _write_input(
        INPUTS / "missing_required_column.xlsx",
        [("MISSING-COLUMN", "MISSING", "", "No model data column exists.")],
        columns=REQUIRED_COLUMNS[:-1],
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "URLs"
    sheet.append(("sku", "Image 1", "Image 2", "Image 3", "Image 4"))
    for scenario in ("VALID-IMAGE", "TRANSPARENT-PNG", "FAIL-THEN-SUCCESS", "HTML-RESPONSE", "PRIVATE-URL"):
        sheet.append((scenario, "", "", "", ""))
        sheet.cell(sheet.max_row, 1).number_format = "@"
    _fit_sheet(sheet, {"sku": 28, "Image 1": 38, "Image 2": 38, "Image 3": 38, "Image 4": 38})
    workbook.save(INPUTS / "image_downloader_uat.xlsx")
    workbook.close()

    images = INPUTS / "images"
    images.mkdir(exist_ok=True)
    Image.new("RGB", (120, 80), "white").save(images / "000123-1.png")
    Image.new("RGBA", (80, 120), (0, 0, 255, 100)).save(images / "ABC-12-2.png")
    Image.new("RGB", (50, 50), "grey").save(images / "ORPHAN-SKU-1.png")
    (images / "malformed.jpg").write_bytes(b"not an image")
    (images / "unsupported_extension.txt").write_text("benign unsupported fixture\n", encoding="utf-8")


def build_ground_truth_template() -> None:
    columns = (
        "attribute set",
        "profile",
        "SKU",
        "base code",
        "source image filenames",
        "attribute header",
        "expected canonical value",
        "expected blank",
        "evidence type",
        "reviewer notes",
        "approval status",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ground Truth"
    sheet.append(columns)
    registry = load_registry(REGISTRY_PATH)
    set_validation = DataValidation(
        type="list", formula1='"' + ",".join(registry.mappings_by_set) + '"'
    )
    blank_validation = DataValidation(type="list", formula1='"TRUE,FALSE"')
    evidence_validation = DataValidation(
        type="list", formula1='"INPUT,TEXT,LABEL,IMAGE,BUSINESS_RULE,NONE"'
    )
    approval_validation = DataValidation(
        type="list", formula1='"PENDING,APPROVED,REJECTED"'
    )
    for validation in (set_validation, blank_validation, evidence_validation, approval_validation):
        sheet.add_data_validation(validation)
    set_validation.add("A2:A1000")
    blank_validation.add("H2:H1000")
    evidence_validation.add("I2:I1000")
    approval_validation.add("K2:K1000")
    _fit_sheet(sheet, {column: 28 for column in columns})
    workbook.save(UAT / "real_product_ground_truth_template.xlsx")
    workbook.close()


def build_expected_headers() -> None:
    registry = load_registry(REGISTRY_PATH)
    payload = {
        "contract_version": "PLAN.md Appendix A",
        "registry_fingerprint": registry.fingerprint,
        "attribute_sets": [
            {
                "canonical_id": set_id,
                "display_name": next(
                    row.attribute_set_name
                    for row in registry.attribute_sets
                    if row.attribute_set_id == set_id
                ),
                "header_count": len(headers),
                "ordered_headers": list(headers),
            }
            for set_id, headers in registry.mappings_by_set.items()
        ],
    }
    (UAT / "expected_headers.json").write_text(
        json.dumps(payload, indent=2) + "\n", encoding="utf-8"
    )


def main() -> None:
    INPUTS.mkdir(parents=True, exist_ok=True)
    build_checklist()
    build_inputs()
    build_ground_truth_template()
    build_expected_headers()
    print("UAT pack regenerated.")


if __name__ == "__main__":
    main()
